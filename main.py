import asyncio
from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.exceptions import RequestValidationError
from fastapi_mail import FastMail, MessageSchema, ConnectionConfig
from sqlalchemy.orm import Session
from sqlalchemy import create_engine, text
from pydantic import BaseModel, EmailStr, validator
import os
import random
import pytz
from datetime import datetime, timedelta
from dotenv import load_dotenv
from fastapi import HTTPException
from typing import Optional
from sqlalchemy import func, and_
from sqlalchemy import desc
from sqlalchemy.orm import aliased
from collections import defaultdict
from dateutil.relativedelta import relativedelta
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, Form, Depends, HTTPException
import pandas as pd
from models import Component, User
from schemas import CompanyRequest
from schemas import ApprovalStatusUpdate
from schemas import SaveTokenRequest
from sqlalchemy import create_engine, MetaData, Table, select
from fastapi.responses import FileResponse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Font, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter



from fastapi.responses import JSONResponse
from fastapi import BackgroundTasks

from models import PanelRiskData 
from models import UPSData
from models import ACData
from models import BatteryLiveData
from models import SwitchboardLiveData
from models import WiringEquipmentData

from schemas import ACRiskResponse

import socket  
import requests
from typing import List
import math
from math import exp


#Absolute imports
from database import SessionLocal, engine
import models, schemas
from schemas import SensorData as SensorDataSchema
from models import User

from passlib.context import CryptContext

ist = pytz.timezone('Asia/Kolkata')

# -------------------- Password Hashing ------------------------
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
# pwd_context = CryptContext(schemes=["argon2"], deprecated="auto")


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


# Load env variables
load_dotenv()

# Create DB tables
models.Base.metadata.create_all(bind=engine)

# FastAPI app
app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, restrict to your frontend domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Dependency for DB session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()



#  -------------------- Email Configuration for sending an email --------------------
conf = ConnectionConfig(
    MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
    MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
    MAIL_FROM=os.getenv("MAIL_FROM"),
    MAIL_PORT=int(os.getenv("MAIL_PORT")),
    MAIL_SERVER=os.getenv("MAIL_SERVER"),
    MAIL_STARTTLS=os.getenv("MAIL_STARTTLS") == "True",
    MAIL_SSL_TLS=os.getenv("MAIL_SSL_TLS") == "True",
    USE_CREDENTIALS=True,
)

#Used to fetch data by ID (given in body)
class SensorIdRequest(BaseModel):
    id: int

class UserCreate(BaseModel):
    name: str
    last_name: str    
    email: EmailStr
    password: str
    confirm_password: str    
    role: Optional[str] = "user"
    assigned_admin: Optional[str] = None
    company_name: str
    # token_number: str


class MessageResponse(BaseModel):
    message: str    

class LoginRequest(BaseModel):
    email: EmailStr
    password: str    

class VerifyOTPRequest(BaseModel):
    email: str
    otp: str 

class ResetPasswordRequest(BaseModel):
    email: str
    password: str 
    confirm_password: str    

class ResendOTPRequest(BaseModel):
    email: EmailStr 

class PanelNameRequest(BaseModel):
    panel_name: str      



#Request schema to pass ups_id or ups_name
class UPSIdRequest(BaseModel):
    ups_id: str = None
    ups_name: str = None  


# -------------------- API Endpoints for sensor data (dummy just for checking remove after development) ----------------------------------------------
@app.post("/sensor-data")
def create_sensor_data(data: schemas.SensorDataCreate, db: Session = Depends(get_db)):
    new_entry = models.SensorData(
        temperature=data.temperature,
        humidity=data.humidity,
        smoke_level=data.smoke_level,
        voltage=data.voltage,
        air_velocity=data.air_velocity,
        pressure=data.pressure,
        timestamp=data.timestamp,
    )
    db.add(new_entry)
    db.commit()
    db.refresh(new_entry)
    return {"message": "Sensor data saved", "id": new_entry.id}
   
         
# -------------------- API Endpoints for fetching the sensor data using ID (dummy just for checking remove after development) ----------------------------------------------
@app.post("/get-sensor-data-by-id", response_model=schemas.SensorData)
def get_sensor_data_by_id_body(request: SensorIdRequest, db: Session = Depends(get_db)):
    data = db.query(models.SensorData).filter(models.SensorData.id == request.id).first()
    if data is None:
        raise HTTPException(status_code=404, detail="Sensor data not found")
    return data


# -------------------- User Registration --------------------
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request, exc):
    errors = [f"{err['loc'][-1]}: {err['msg']}" for err in exc.errors()]
    return JSONResponse(status_code=400, content={"message": ", ".join(errors)})


# -----------------------------------------------------------------------------------------------------
# Endpoint to register a new user.
# Accepts user details, validates them, and stores in the database.
# @app.post("/register", response_model=schemas.MessageResponse)
# async def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):

#     # Check if email already exists
#     if db.query(models.User).filter(models.User.email == user.email).first():
#         return JSONResponse(status_code=400, content={"message": "Email already registered"})

#     # Check if contact number already exists
#     if db.query(models.User).filter(models.User.contact_number == user.contact_number).first():
#         return JSONResponse(status_code=400, content={"message": "Contact number already registered"})

#     # Hash the password
#     hashed_password = get_password_hash(user.password)
#     hashed_confirm_password = get_password_hash(user.confirm_password)

#     # Timezone
#     india = pytz.timezone("Asia/Kolkata")

#     # Generate OTP
#     otp = str(random.randint(100000, 999999))
#     otp_expiry = datetime.now(india) + timedelta(minutes=10)

#     # Create user
#     new_user = models.User(
#         name=user.name,
#         last_name=user.last_name,
#         email=user.email,
#         contact_number=user.contact_number,
#         password=hashed_password,
#         confirm_password=hashed_confirm_password,
#         otp=otp,
#         otp_expiry=otp_expiry,
#         otp_status=0
#     )

#     # Save to DB
#     db.add(new_user)
#     db.commit()
#     db.refresh(new_user)

#     # Send confirmation email
#     message = MessageSchema(
#         subject="Your OTP for Registration",
#         recipients=[user.email],
#         body=f"Hello {user.name},\n\nYour OTP for completing registration is: {otp}\n\nThis OTP is valid for 10 minutes.\n\nBest regards,\nMepstra IT Solutions",
#         subtype="plain",
#     )
#     fm = FastMail(conf)
#     await fm.send_message(message)

#     return {"message": "Registration successful"}


# @app.post("/register", response_model=schemas.MessageResponse)
# async def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):

#     # Check if email already exists
#     if db.query(models.User).filter(models.User.email == user.email).first():
#         return JSONResponse(status_code=400, content={"message": "Email already registered"})

#     # Check if contact number already exists
#     if db.query(models.User).filter(models.User.contact_number == user.contact_number).first():
#         return JSONResponse(status_code=400, content={"message": "Contact number already registered"})

#     # Hash the password
#     hashed_password = get_password_hash(user.password)
#     hashed_confirm_password = get_password_hash(user.confirm_password)

#     # Timezone
#     india = pytz.timezone("Asia/Kolkata")

#     # Generate OTP
#     otp = str(random.randint(100000, 999999))
#     otp_expiry = datetime.now(india) + timedelta(minutes=10)

#     # Create user with new columns
#     new_user = models.User(
#         name=user.name,
#         last_name=user.last_name,
#         email=user.email,
#         contact_number=user.contact_number,
#         password=hashed_password,
#         confirm_password=hashed_confirm_password,
#         otp=otp,
#         otp_expiry=otp_expiry,
#         otp_status=0,   
#         role=user.role,           
#         assigned_admin=user.assigned_admin
#     )

#     db.add(new_user)
#     db.commit()
#     db.refresh(new_user)

#     # Send confirmation email
#     message = MessageSchema(
#         subject="Your OTP for Registration",
#         recipients=[user.email],
#         body=f"Hello {user.name},\n\nYour OTP for completing registration is: {otp}\n\nThis OTP is valid for 10 minutes.\n\nBest regards,\nMepstra IT Solutions",
#         subtype="plain",
#     )
#     fm = FastMail(conf)
#     await fm.send_message(message)

#     return {"message": "Registration successful"}


# @app.post("/register", response_model=schemas.MessageResponse)
# async def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
 
#     # ✅ Check company name mandatory (not empty / just spaces)
#     if not user.company_name or not user.company_name.strip():
#         return JSONResponse(
#             status_code=400,
#             content={"message": "Company name is required"}
#         )
 
#     # Check if email already exists
#     if db.query(models.User).filter(models.User.email == user.email).first():
#         return JSONResponse(
#             status_code=400,
#             content={"message": "Email already registered"}
#         )
 
#     # Check if contact number already exists
#     if db.query(models.User).filter(models.User.contact_number == user.contact_number).first():
#         return JSONResponse(
#             status_code=400,
#             content={"message": "Contact number already registered"}
#         )
 
#     # Hash the password
#     hashed_password = get_password_hash(user.password)
#     hashed_confirm_password = get_password_hash(user.confirm_password)
 
#     # Timezone
#     india = pytz.timezone("Asia/Kolkata")
 
#     # Generate OTP
#     otp = str(random.randint(100000, 999999))
#     otp_expiry = datetime.now(india) + timedelta(minutes=10)
 
#     # Create user with new columns
#     new_user = models.User(
#         name=user.name,
#         last_name=user.last_name,
#         email=user.email,
#         contact_number=user.contact_number,
#         password=hashed_password,
#         confirm_password=hashed_confirm_password,
#         company_name=user.company_name,  
#         otp=otp,
#         otp_expiry=otp_expiry,
#         otp_status=0,
#         role=user.role,
#         assigned_admin=user.assigned_admin
#     )
 
#     db.add(new_user)
#     db.commit()
#     db.refresh(new_user)
 
#     # Send confirmation email
#     message = MessageSchema(
#         subject="Your OTP for Registration",
#         recipients=[user.email],
#         body=(
#             f"Hello {user.name},\n\n"
#             f"Your OTP for completing registration is: {otp}\n\n"
#             f"This OTP is valid for 10 minutes.\n\n"
#             f"Best regards,\nMepstra IT Solutions"
#         ),
#         subtype="plain",
#     )
#     fm = FastMail(conf)
#     await fm.send_message(message)
 
#     return {"message": "Registration successful"}



@app.post("/register", response_model=schemas.MessageResponse)
async def register_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
 
    # Company name mandatory
    if not user.company_name or not user.company_name.strip():
        return JSONResponse(
            status_code=400,
            content={"message": "Company name is required"}
        )
 
    # Email check
    if db.query(models.User).filter(models.User.email == user.email).first():
        return JSONResponse(
            status_code=400,
            content={"message": "Email already registered"}
        )
 
    # Contact number check
    if db.query(models.User).filter(models.User.contact_number == user.contact_number).first():
        return JSONResponse(
            status_code=400,
            content={"message": "Contact number already registered"}
        )
 
    # 👉 NEW: Allow only 1 admin per company
    if user.role.lower() == "admin":
        existing_admin = (
            db.query(models.User)
            .filter(models.User.company_name == user.company_name)
            .filter(models.User.role == "admin")
            .first()
        )

        if existing_admin:
            return JSONResponse(
                status_code=400,
                content={
                    "message": f"Admin already registered for company '{user.company_name}'"
                }
            )
       
 
    # Hash password
    hashed_password = get_password_hash(user.password)
    hashed_confirm_password = get_password_hash(user.confirm_password)
 
    india = pytz.timezone("Asia/Kolkata")
 
    otp = str(random.randint(100000, 999999))
    otp_expiry = datetime.now(india) + timedelta(minutes=10)

    #approval_status logic
    if user.role.lower() == "admin":
        approval_status = "Accepted"
    else:
        approval_status = "Pending"
 
    new_user = models.User(
        name=user.name,
        last_name=user.last_name,
        email=user.email,
        contact_number=user.contact_number,
        password=hashed_password,
        confirm_password=hashed_confirm_password,
        company_name=user.company_name,
        otp=otp,
        otp_expiry=otp_expiry,
        otp_status=0,
        role=user.role,
        assigned_admin=user.assigned_admin,
        approval_status=approval_status

    )
 
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
 
    message = MessageSchema(
        subject="Your OTP for Registration",
        recipients=[user.email],
        body=(
            f"Hello {user.name},\n\n"
            f"Your OTP for completing registration is: {otp}\n\n"
            f"This OTP is valid for 10 minutes.\n\n"
            f"Best regards,\nMepstra IT Solutions"
        ),
        subtype="plain",
    )
    fm = FastMail(conf)
    await fm.send_message(message)
 
    return {"message": "Registration successful"}




#-----------------------------------------------------------------------------------------------------
# Endpoint to verify the otp entered by user.
# Verifies the entered otp with the database otp if matches then gives message OTP verified successfully.
@app.post("/verify-otp")
async def verify_otp(request: VerifyOTPRequest, db: Session = Depends(get_db)):
    email = request.email
    otp = request.otp

    # 1️⃣ Find user by email
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # 2️⃣ Check if OTP matches
    if user.otp != otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")

    # 3️⃣ Check if OTP expired (IST)
    if not user.otp_expiry:
        raise HTTPException(status_code=400, detail="OTP not generated")

    india = pytz.timezone("Asia/Kolkata")
    
    now_ist = datetime.now(india).replace(tzinfo=None)

    if now_ist > user.otp_expiry:
        raise HTTPException(status_code=400, detail="OTP expired")

    # 4️⃣ Mark OTP as verified
    user.otp_status = True
    db.commit()
    db.refresh(user)

    return {"message": "OTP verified successfully"}



#-----------------------------------------------------------------------------------------------------
# Endpoint to resend the otp for user verification.
# Sends an new otp and stored in database for verification.
@app.post("/resend-otp")
async def resend_otp(request: ResendOTPRequest, db: Session = Depends(get_db)):
    email = request.email

    #Find user
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    #Generate new OTP
    otp = str(random.randint(100000, 999999))
    india = pytz.timezone("Asia/Kolkata")
    otp_expiry = datetime.now(india) + timedelta(minutes=10)

    #Update user
    user.otp = otp
    user.otp_expiry = otp_expiry
    user.otp_status = False
    db.commit()
    db.refresh(user)

    #Send OTP email
    message = MessageSchema(
        subject="Your New OTP for Verification",
        recipients=[user.email],
        body=f"Hello {user.name},\n\nYour new OTP is: {otp}\n\nThis OTP is valid for 10 minutes.\n\nBest regards,\nMepstra IT Solutions",
        subtype="plain",
    )
    fm = FastMail(conf)
    await fm.send_message(message)

    return {"message": "New OTP sent successfully"}


#---------------------------------Duplicate Api remove after development--------------------------------------------------------------------
# Endpoint to resend the otp for user verification.
# @app.post("/resend-otp")
# async def resend_otp(request: ResendOTPRequest, db: Session = Depends(get_db)):
#     email = request.email

#     #Find user
#     user = db.query(models.User).filter(models.User.email == email).first()
#     if not user:
#         raise HTTPException(status_code=404, detail="User not found")

#     #Generate new OTP
#     otp = str(random.randint(100000, 999999))
#     india = pytz.timezone("Asia/Kolkata")
#     otp_expiry = datetime.now(india) + timedelta(minutes=10)

#     #Update user
#     user.otp = otp
#     user.otp_expiry = otp_expiry
#     user.otp_status = False
#     db.commit()
#     db.refresh(user)

#     #Send OTP email
#     message = MessageSchema(
#         subject="Your New OTP for Verification",
#         recipients=[user.email],
#         body=f"Hello {user.name},\n\nYour new OTP is: {otp}\n\nThis OTP is valid for 10 minutes.\n\nBest regards,\nMepstra IT Solutions",
#         subtype="plain",
#     )
#     fm = FastMail(conf)
#     await fm.send_message(message)

#     return {"message": "New OTP sent successfully"}




#--------------------------Sample Testing sending Email(Dummy just for checking remove after development)---------------------------------
@app.get("/test-email")
async def test_email():
    message = MessageSchema(
        subject="Test Email",
        recipients=[os.getenv("MAIL_USERNAME")],
        body="This is a test email from Fire Prediction App",
        subtype="plain",
    )
    fm = FastMail(conf)
    await fm.send_message(message)
    return {"message": "Test email sent successfully!"}


#---------------------------------------------------------------------------------------------------------------
# Endpoint to Login theuser.
# Verifies user details, validates them, allows only if the entered user name and password matches from database.
# @app.post("/login")
# def login(request: LoginRequest, db: Session = Depends(get_db)):
#     # find user by email
#     user = db.query(User).filter(User.email == request.email).first()
#     if not user:
#         raise HTTPException(status_code=401, detail="Invalid email or password")
    
#     if not pwd_context.verify(request.password, user.password):
#         raise HTTPException(status_code=401, detail="Invalid email or password")


#     # # verify password
#     # if not pwd_context.verify(request.password, user.password):
#     #     raise HTTPException(status_code=401, detail="Invalid email or password")

#     # check OTP verification status
#     if not user.otp_status:
#         raise HTTPException(
#             status_code=403,
#             detail="Please verify your OTP before logging in"
#         )
#     return {"message": "Login successful"}



# @app.post("/login")
# def login(request: LoginRequest, db: Session = Depends(get_db)):
#     # 🔹 Find user by email
#     user = db.query(User).filter(User.email == request.email).first()

#     if not user:
#         raise HTTPException(status_code=401, detail="Invalid email or password")

#     # 🔹 Verify password
#     if not pwd_context.verify(request.password, user.password):
#         raise HTTPException(status_code=401, detail="Invalid email or password")

#     # 🔹 Check OTP verification status
#     if not user.otp_status:
#         raise HTTPException(
#             status_code=403,
#             detail="Please verify your OTP before logging in"
#         )

#     # 🔹 Return user info (including role and assigned admin)
#     return {
#         "message": "Login successful",
#         "user": {
#             "name": user.name,
#             "email": user.email,
#             "role": user.role,
#             "assigned_admin": user.assigned_admin,
#             "company_name": user.company_name
#         }
#     }    


@app.post("/login")
def login(request: LoginRequest, db: Session = Depends(get_db)):
    # 🔹 Find user by email
    user = db.query(User).filter(User.email == request.email).first()

    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")

    # 🔹 Verify password
    if not pwd_context.verify(request.password, user.password):
        raise HTTPException(status_code=401, detail="Invalid email or password")

    # 🔹 Check OTP verification status
    if not user.otp_status:
        raise HTTPException(
            status_code=403,
            detail="Please verify your OTP before logging in"
        )

    # 🔹 Check approval status
    if user.approval_status == "Pending":
        raise HTTPException(
            status_code=403,
            detail="Your approval is still pending. Please wait for the admin to verify your details"
        )

    if user.approval_status == "Rejected":
        raise HTTPException(
            status_code=403,
            detail="Your registration details has been rejected by the admin."
        )

    # 🔹 If approval_status is Accepted → login success
    return {
        "message": "Login successful",
        "user": {
            "name": user.name,
            "email": user.email,
            "role": user.role,
            "assigned_admin": user.assigned_admin,
            "company_name": user.company_name,
            
        }
    }

#-----------------------------------------------------------------------------------------------------
# Endpoint to forgot-password for the user.
# If user forgots password then he can do forgot-password and reset the new password after successful verification of user.
@app.post("/forgot-password")
async def forgot_password(email: str, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Generate OTP
    otp = str(random.randint(100000, 999999))

    # Set OTP expiry (10 mins from now IST)
    india = pytz.timezone("Asia/Kolkata")
    otp_expiry = datetime.now(india).replace(tzinfo=None) + timedelta(minutes=10)

    user.otp = otp
    user.otp_expiry = otp_expiry
    user.otp_status = False
    db.commit()
    db.refresh(user)

    # Send OTP via email
    message = MessageSchema(
        subject="Password Reset OTP",
        recipients=[email],
        body=f"Hello {user.name},\n\nYour OTP for password reset is: {otp}\nIt is valid for 10 minutes.\n\nBest regards,\nMepstra IT Solutions",
        subtype="plain"
    )
    fm = FastMail(conf)
    await fm.send_message(message)

    return {"message": "OTP sent to your email"} 


#-----------------------------------------------------------------------------------------------------
# Endpoint to verify-forgot-otp for the user.
# If user clicks forgots password then he must verify with otp if entered otp matches then only he will allow to change the password.
@app.post("/verify-forgot-otp")
def verify_forgot_otp(request: VerifyOTPRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == request.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    india = pytz.timezone("Asia/Kolkata")
    now_ist = datetime.now(india).replace(tzinfo=None)

    if user.otp != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    if not user.otp_expiry or now_ist > user.otp_expiry:
        raise HTTPException(status_code=400, detail="OTP expired")

    # Mark OTP verified
    user.otp_status = True
    db.commit()
    db.refresh(user)

    return {"message": "OTP verified successfully"}   



#-----------------------------------------------------------------------------------------------------
# Endpoint to reset-password(change password) for the user.
# If user wants to change his password he can enter the email with new password and after verification with otp the new password will be updated.
@app.post("/reset-password")
def reset_password(
    request: ResetPasswordRequest, 
    db: Session = Depends(get_db), 
    background_tasks: BackgroundTasks = None
):
    user = db.query(models.User).filter(models.User.email == request.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    if not user.otp_status:
        raise HTTPException(status_code=403, detail="Please verify OTP first")

    if request.password != request.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")

    # Hash new password
    hashed_password = pwd_context.hash(request.password)
    user.password = hashed_password

    # Reset OTP fields
    user.otp_status = False
    user.otp = None
    user.otp_expiry = None

    db.commit()
    db.refresh(user)

    # Prepare email
    message = MessageSchema(
        subject="Password Reset Successful",
        recipients=[user.email],
        body=(
            f"Hello {user.name},\n\n"
            "Your password has been changed successfully.\n"
            "If you did not perform this action, please contact support immediately.\n\n"
            "Best regards,\nMepstra IT Solutions."
        ),
        subtype="plain"
    )

    # Send email in background (optional)
    if background_tasks:
        try:
            background_tasks.add_task(FastMail(conf).send_message, message)
        except Exception as e:
            # Log instead of failing the API
            print(f"[Warning] Failed to send email: {e}")

    return {"message": "Password reset successfully. Email notification (if possible) sent."}



#-----------------------------------------------------------------------------------------------------
# Endpoint to get user details to display all details in profile.
# With the user id fetching the details from database and sending in response.
@app.post("/get-user-profile", response_model=schemas.UserProfileResponse)
def get_user_profile(request: schemas.EmailRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == request.email).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    return {
        "name": user.name,
        "last_name": user.last_name,
        "contact_number": user.contact_number,
        "email": user.email,
        "company_name": user.company_name
    }




#-----------------------------------------------------------------------------------------------------
# Endpoint to update-user-profile details.
# User can update his details from profile click update and saved in database.
@app.post("/update-user-profile", response_model=schemas.UserProfileResponse)
def update_user_profile(request: schemas.UpdateUserRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == request.email).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Update only provided fields
    if request.name is not None:
        user.name = request.name
    if request.last_name is not None:
        user.last_name = request.last_name
    if request.contact_number is not None:
        user.contact_number = request.contact_number

    db.commit()
    db.refresh(user)

    return {
        "name": user.name,
        "last_name": user.last_name,
        "contact_number": user.contact_number,
        "email": user.email
    }





# Function to calculate_risk is to calculate the panel risk using the data from database.
#-----------------------------------------------------------------------------------------------------
# Endpoint to calculate_risk of panels.
# We are fetching the data from the panel_risk_data and Using them in algorithm and formula and predicting the risk level.
def calculate_risk(data):
    # Prevent division by zero
    I_abn = 0 if data.rated_current == 0 else abs(data.I_phaseA - data.rated_current) / data.rated_current

    # Risk sub-scores
    CRS = 0.5 * data.R_contact_norm + 0.3 * data.T_contact_norm + 0.2 * I_abn             #CRS - Contact Risk Score
    PDS = 0.6 * data.HF_I_env_norm + 0.4 * data.US_sig_norm                               #PDS - Power Distribution Score 
    ATS = 0.5 * data.transient_peak_norm + 0.3 * data.pulse_rate_norm + 0.2 * data.HF_energy_norm  #ATS – Arc/Transient Score
    TRS = 0.5 * data.T_contact_norm + 0.3 * data.thermal_gradient_norm + 0.2 * data.T_bus_norm     #TRS – Thermal/Temperature Risk Score
    IDS = 0.6 * (1 - data.IR_norm) + 0.4 * data.PD_trend_norm                                      #IDS – Insulation/Dielectric Score
    LFS = 0.7 * data.I_leak_norm + 0.3 * data.neutral_imbalance                                    #LFS – Leakage/Load Flow Score  

    PRS = 0.2 * CRS + 0.15 * PDS + 0.25 * ATS + 0.2 * TRS + 0.1 * IDS + 0.1 * LFS

    # Determine risk level
    if PRS < 0.5:
        risk_level = "Normal"
    elif PRS < 0.7:
        risk_level = "Warning"
    elif PRS < 0.85:
        risk_level = "High"
    else:
        risk_level = "High"

    return PRS, risk_level

@app.post("/calculate_risk", response_model=schemas.RiskResponse)
def create_panel_risk(request: PanelNameRequest, db: Session = Depends(get_db)):
    panel = db.query(models.PanelRiskData).filter(models.PanelRiskData.panel_name == request.panel_name).first()
    
    if not panel:
        raise HTTPException(status_code=404, detail="Panel not found")
    
    PRS, risk_level = calculate_risk(panel)
    panel.risk_score = PRS
    panel.risk_level = risk_level
    db.commit()
    db.refresh(panel)

    return schemas.RiskResponse(
        risk_score=panel.risk_score,
        risk_level=panel.risk_level,
        created_at=panel.created_at.isoformat()
    )


@app.post("/calculate_risk_all_panel", response_model=dict)
def calculate_risk_all(db: Session = Depends(get_db)):
    # Fetch all panel records
    panels = db.query(models.PanelRiskData).all()

    if not panels:
        raise HTTPException(status_code=404, detail="No panel data found")

    updated_panels = []

    for panel in panels:
        # Calculate risk for each row
        PRS, risk_level = calculate_risk(panel)

        # Update table fields
        panel.risk_score = PRS
        panel.risk_level = risk_level

        updated_panels.append({
            "instance_id": panel.instance_id,
            "risk_score": round(PRS, 4),
            "risk_level": risk_level
        })

    # Commit all updates once
    db.commit()

    print("Panel risk calculation running...")
    return {
        "message": f"Risk calculation completed for {len(updated_panels)} panels.",
        "updated_data": updated_panels
    }


#---------------------------------------------------------------------------------------------------------
# Risk calculation function
def real_values_calculate_risk(panel):
    # Prevent division by zero
    I_abn = abs(panel.I_phaseA - panel.rated_current) / panel.rated_current if panel.rated_current else 0

    # Risk sub-scores
    CRS = 0.5 * panel.R_contact_norm + 0.3 * panel.T_contact_norm + 0.2 * I_abn  #CRS - Contact Risk Score
    PDS = 0.6 * panel.HF_I_env_norm + 0.4 * panel.US_sig_norm                    #PDS - Power Distribution Score 
    ATS = 0.5 * panel.transient_peak_norm + 0.3 * panel.pulse_rate_norm + 0.2 * panel.HF_energy_norm  #ATS – Arc/Transient Score
    TRS = 0.5 * panel.T_contact_norm + 0.3 * panel.thermal_gradient_norm + 0.2 * panel.T_bus_norm    #TRS – Thermal/Temperature Risk Score
    IDS = 0.6 * (1 - panel.IR_norm) + 0.4 * panel.PD_trend_norm                                    #IDS – Insulation/Dielectric Score
    LFS = 0.7 * panel.I_leak_norm + 0.3 * panel.neutral_imbalance                                    #LFS – Leakage/Load Flow Score  

    PRS = 0.2 * CRS + 0.15 * PDS + 0.25 * ATS + 0.2 * TRS + 0.1 * IDS + 0.1 * LFS

    # Determine risk level
    if PRS < 0.5:
        risk_level = "Normal"
    elif PRS < 0.7:
        risk_level = "Warning"
    elif PRS < 0.85:
        risk_level = "High"
    else:
        risk_level = "High"

    return PRS, risk_level

#Request schema to only pass panel_name
class PanelNameRequest(schemas.BaseModel):
    panel_name: str

@app.post("/real_value_update_panel_risk", response_model=schemas.RiskResponse)
def update_panel_risk(request: PanelNameRequest, db: Session = Depends(get_db)):
    # Fetch the panel from DB
    panel = db.query(models.PanelRiskData).filter(models.PanelRiskData.panel_name == request.panel_name).first()
    if not panel:
        raise HTTPException(status_code=404, detail="Panel not found")

    # Calculate updated risk
    PRS, risk_level = real_values_calculate_risk(panel)

    # Update only risk_score and risk_level
    panel.risk_score = PRS
    panel.risk_level = risk_level
    db.commit()
    db.refresh(panel)

    # Return response
    return schemas.RiskResponse(
        risk_score=panel.risk_score,
        risk_level=panel.risk_level,
        created_at=panel.created_at.isoformat()
    )



#Request body model
class UPSIDRequest(BaseModel):
    instance_id: str

#TRS calculation
def calculate_trs(ups: UPSData):

    # Safe max values for normalization
    MAX_VALUES = {
        "I": 100.0,
        "TUPS": 80.0,
        "IL": 30.0,
        "THD": 15.0
    }

    # Weight factors
    WEIGHTS = {
        "I": 0.2,
        "V": 0.0,  # Not used
        "PF": 0.1,
        "T": 0.3,
        "IL": 0.2,
        "THD": 0.1
    }
    I_score = ups.I / MAX_VALUES["I"]
    T_score = ups.TUPS / MAX_VALUES["TUPS"]
    IL_score = ups.IL / MAX_VALUES["IL"]
    THD_score = ups.THD / MAX_VALUES["THD"]
    PF_score = 1 - ups.PF

    TRS = (
        WEIGHTS["I"] * I_score +
        WEIGHTS["V"] * 0 +
        WEIGHTS["PF"] * PF_score +
        WEIGHTS["T"] * T_score +
        WEIGHTS["IL"] * IL_score +
        WEIGHTS["THD"] * THD_score
    )

    if TRS < 0.3:
        risk_level = "Normal"
    elif TRS < 0.6:
        risk_level = "Medium"
    else:
        risk_level = "High"

    TRS = round(TRS * 100, 2)
    TRS = min(TRS, 100)        

    return TRS, risk_level

@app.post("/calculate_ups_risk", response_model=schemas.RiskResponse)
def calculate_ups_risk(request: UPSIDRequest, db: Session = Depends(get_db)):
    # Fetch UPS live data
    ups = db.query(UPSData).filter(UPSData.instance_id == request.instance_id).first()
    if not ups:
        raise HTTPException(status_code=404, detail="instance_id not found")

    # Calculate TRS and risk
    trs, risk_level = calculate_trs(ups)

    # Update DB
    ups.risk_score = round(trs, 3)
    ups.risk_level = risk_level
    ups.risk_created_at = datetime.utcnow()
    db.commit()
    db.refresh(ups)


    print("UPS risk calculation running...")
    
    # Return response
    return schemas.RiskResponse(
        risk_score=ups.risk_score,
        risk_level=ups.risk_level,
        created_at=ups.risk_created_at.isoformat()
    )



#TRS calculation logic (unchanged)
def calculate_trs(ups: models.UPSData):

    MAX_VALUES = {
        "I": 100.0,
        "TUPS": 80.0,
        "IL": 30.0,
        "THD": 15.0
    }

    WEIGHTS = {
        "I": 0.2,
        "V": 0.0,
        "PF": 0.1,
        "T": 0.3,
        "IL": 0.2,
        "THD": 0.1
    }

    I_score = ups.I / MAX_VALUES["I"]
    T_score = ups.TUPS / MAX_VALUES["TUPS"]
    IL_score = ups.IL / MAX_VALUES["IL"]
    THD_score = ups.THD / MAX_VALUES["THD"]
    PF_score = 1 - ups.PF

    TRS = (
        WEIGHTS["I"] * I_score +
        WEIGHTS["V"] * 0 +
        WEIGHTS["PF"] * PF_score +
        WEIGHTS["T"] * T_score +
        WEIGHTS["IL"] * IL_score +
        WEIGHTS["THD"] * THD_score
    )

    if TRS < 0.3:
        risk_level = "Normal"
    elif TRS < 0.6:
        risk_level = "Medium"
    else:
        risk_level = "High"

    TRS = round(TRS * 100, 2)
    TRS = min(TRS, 100)    

    return TRS, risk_level


#New endpoint: calculate risk for all UPS instances
@app.post("/calculate_all_ups_risk")
def calculate_all_ups_risk(db: Session = Depends(get_db)):
    # Fetch all UPS instances
    ups_records = db.query(models.UPSData).all()

    if not ups_records:
        raise HTTPException(status_code=404, detail="No UPS records found")

    results = []

    for ups in ups_records:
        trs, risk_level = calculate_trs(ups)

        # Update record
        ups.risk_score = round(trs, 3)
        ups.risk_level = risk_level
        ups.risk_created_at = datetime.utcnow()

        results.append({
            "instance_id": ups.instance_id,
            "risk_score": ups.risk_score,
            "risk_level": ups.risk_level,
            "updated_at": ups.risk_created_at.isoformat()
        })

    db.commit()

    print("UPS risk calculation running...")

    return {
        "message": "TRS and risk levels calculated for all UPS instances.",
        "total_updated": len(results),
        "details": results
    }

# ----------------------------------------------------------------------------------------------------
class ACIDRequest(BaseModel):
    instance_id: str

#FARI calculation function
def calculate_fari(ac_data: ACData):
    # Constants
    α1, α2 = 0.7, 0.3
    β1, β2 = 0.6, 0.4
    W = [2.0, 1.5, 1.2, 1.0, 1.3]  # weights

    # Derived indices
    EAI = α1 * abs(ac_data.I_measured - ac_data.Irated)/ac_data.Irated + α2 * (ac_data.THD / ac_data.THDmax)
    TSI = (ac_data.Tc - ac_data.Tsafe) / ac_data.Tsafe
    CHI = β1 * (ac_data.ESR / ac_data.ESRlimit) + β2 * ac_data.C_drop
    MSI = ac_data.RHpcb / ac_data.RHthreshold
    CIRI = (ac_data.Ires / ac_data.Ilimit) + (1 / ac_data.IR)

    # Weighted FARI
    FARIraw = W[0]*EAI + W[1]*TSI + W[2]*CHI + W[3]*MSI + W[4]*CIRI
    FARInorm = FARIraw / sum(W)

    # Risk classification
    if FARInorm < 0.5:
        risk_level = "Low"
    elif 0.5 <= FARInorm < 1.0:
        risk_level = "Medium"
    elif 1.0 <= FARInorm < 1.5:
        risk_level = "High"
    else:
        risk_level = "High"


    
    return FARInorm, risk_level

#API endpoint
@app.post("/calculate_ac_risk", response_model=ACRiskResponse)
def calculate_ac_risk(request: ACIDRequest, db: Session = Depends(get_db)):
    # Fetch AC data
    ac_data = db.query(ACData).filter(ACData.instance_id == request.instance_id).first()
    if not ac_data:
        raise HTTPException(status_code=404, detail="AC data not found")

    # Calculate FARI and risk
    fari_score, risk_level = calculate_fari(ac_data)
    
    fari_score = round(fari_score * 100, 2)
    # Update DB
    ac_data.risk_score = round(fari_score, 3)
    ac_data.risk_level = risk_level
    ac_data.risk_created_at = datetime.utcnow()
    db.commit()
    db.refresh(ac_data)


    # Return response
    return ACRiskResponse(
        risk_score=ac_data.risk_score,
        risk_level=ac_data.risk_level,
        created_at=ac_data.risk_created_at.isoformat()
    )


# Request model
class BatteryIDRequest(BaseModel):
    instance_id: str

# Response model
class BatteryRiskResponse(BaseModel):
    instance_id: str
    risk_score: float
    risk_level: str
    created_at: str

# Calculation function (simple example)
def calculate_battery_risk(data: BatteryLiveData):
    """
    Example risk calculation based on live battery parameters.
    Replace with your actual model/formula later.
    """

    # Normalize values for scoring
    I_factor = abs(data.I_meas) / 100.0
    V_factor = abs(data.V_meas - 48) / 48.0      # assume 48V nominal
    T_factor = data.T_meas / 100.0
    SOC_factor = (100 - data.SOC_pred) / 100.0
    Cycle_factor = data.Cycle_pred / 1000.0
    IR_factor = data.IR_pred / 0.05              # assume 0.05Ω limit

    # Weighted risk score
    raw_score = (0.2 * I_factor +
                 0.2 * V_factor +
                 0.2 * T_factor +
                 0.2 * SOC_factor +
                 0.1 * Cycle_factor +
                 0.1 * IR_factor)

    risk_score = round(raw_score, 3)

    # if risk_score <= 0.3:
    #     risk_level = "Normal"
    # elif 0.3 <= risk_score < 0.6:
    #     risk_level = "Medium"
    # elif 0.6 <= risk_score < 0.8:
    #     risk_level = "High"
    # else:  # risk_score >= 0.8
    #     risk_level = "High"    

    if risk_score < 0.3:
        risk_level = "Normal"
    elif 0.3 <= risk_score < 0.6:
        risk_level = "Medium"
    elif risk_score >= 0.6:
        risk_level = "High"

    # Convert to percentage (e.g., 0.45 → 45.0%)
    risk_score = round(risk_score * 100, 2)
    risk_score = min(risk_score, 100)

    return risk_score, risk_level

# API endpoint
@app.post("/calculate_battery_risk", response_model=BatteryRiskResponse)
def calculate_battery_risk_api(request: BatteryIDRequest, db: Session = Depends(get_db)):
    # Fetch battery record
    battery_data = db.query(BatteryLiveData).filter(BatteryLiveData.instance_id == request.instance_id).first()
    if not battery_data:
        raise HTTPException(status_code=404, detail="Battery data not found")

    # Run risk calculation
    risk_score, risk_level = calculate_battery_risk(battery_data)

    #Update DB
    battery_data.risk_score = risk_score
    battery_data.risk_level = risk_level
    battery_data.risk_created_at = datetime.utcnow()
    db.commit()
    db.refresh(battery_data)

    # Return response
    return BatteryRiskResponse(
        instance_id=battery_data.instance_id,  #  include this
        risk_score=battery_data.risk_score,
        risk_level=battery_data.risk_level,
        created_at=battery_data.risk_created_at.isoformat()
    )


@app.post("/calculate_all_battery_risk", response_model=List[BatteryRiskResponse])
def calculate_battery_risk_all_api(db: Session = Depends(get_db)):
    # Fetch all battery records
    all_batteries = db.query(BatteryLiveData).all()
    if not all_batteries:
        raise HTTPException(status_code=404, detail="No battery data found")

    responses = []

    for battery_data in all_batteries:
        # Run risk calculation for each record
        risk_score, risk_level = calculate_battery_risk(battery_data)

        # Update DB record
        battery_data.risk_score = risk_score
        battery_data.risk_level = risk_level
        # battery_data.risk_created_at = datetime.utcnow()
        battery_data.risk_created_at = datetime.now()


        responses.append(BatteryRiskResponse(
            instance_id=battery_data.instance_id,
            risk_score=risk_score,
            risk_level=risk_level,
            created_at=battery_data.risk_created_at.isoformat()
            
        ))

    # Commit all updates at once
    db.commit()

    print("Battery risk calculation running...")

    return responses

#------------------------------------------------------------------------------------------------------
class SwitchboardIDRequest(BaseModel):
    instance_id: str

class SwitchboardRiskResponse(BaseModel):
    risk_score: float
    risk_level: str
    created_at: str

#------------------ Calculation Function ------------------ #
def calculate_switchboard_risk(data: SwitchboardLiveData):
    I_RATED = 400.0
    V_NOM = 500.0
    T_SAFE = 80.0
    I_LEAK_MAX = 10.0
    R_REF = 0.05
    E_SAFE = 150.0
    DELTA_T_SAFE = 50.0

    w1, w2, w3, w4, w5, w6, w7, w8 = 0.1, 0.1, 0.1, 0.05, 0.1, 0.05, 0.15, 0.2

    X = (
        w1 * (data.I_load / I_RATED) +
        w2 * (data.V_meas / V_NOM) +
        w3 * (data.T_meas / T_SAFE) +
        w4 * (data.I_leak / I_LEAK_MAX) +
        w5 * (data.R_contact / R_REF) +
        w6 * (data.R_surface / R_REF) +
        w7 * (data.E_arc / E_SAFE) +
        w8 * (data.deltaT_pred / DELTA_T_SAFE)
    )

    # Sigmoid function
    risk_score = 1 / (1 + exp(-X))
    risk_score = round(risk_score, 3)


    if risk_score < 0.3:
        risk_level = "Normal"
    elif 0.3 <= risk_score <= 0.6:
        risk_level = "Medium"
    else:
        risk_level = "High"    

    return risk_score, risk_level

#------------------ API Endpoint ----------------------------------------------
@app.post("/calculate_switchboard_risk", response_model=SwitchboardRiskResponse)
def calculate_switchboard_risk_api(request: SwitchboardIDRequest, db: Session = Depends(get_db)):
    # Fetch the switchboard data
    sb_data = db.query(SwitchboardLiveData).filter(
        SwitchboardLiveData.instance_id == request.instance_id
    ).first()  # <-- MUST use .first() to get a single row

    if not sb_data:
        raise HTTPException(status_code=404, detail="Switchboard data not found")

    # Run calculation
    risk_score, risk_level = calculate_switchboard_risk(sb_data)

    # Update DB
    sb_data.risk_score = risk_score
    sb_data.risk_level = risk_level
    sb_data.risk_created_at = datetime.utcnow()
    db.commit()
    db.refresh(sb_data)

    return SwitchboardRiskResponse(
        risk_score=sb_data.risk_score,
        risk_level=sb_data.risk_level,
        created_at=sb_data.risk_created_at.isoformat()
    )


@app.post("/calculate_all_switchboard_risk", response_model=dict)
def calculate_switchboard_risk_api(db: Session = Depends(get_db)):
    # Fetch all switchboard rows
    all_switchboards = db.query(SwitchboardLiveData).all()

    if not all_switchboards:
        raise HTTPException(status_code=404, detail="No switchboard data found")

    updated_count = 0

    for sb_data in all_switchboards:
        try:
            # Perform risk calculation
            risk_score, risk_level = calculate_switchboard_risk(sb_data)

            # Update database record
            sb_data.risk_score = risk_score
            sb_data.risk_level = risk_level
            sb_data.risk_created_at = datetime.utcnow()
            db.commit()
            db.refresh(sb_data)
            updated_count += 1
        except Exception as e:
            print(f"Error updating instance_id {sb_data.instance_id}: {e}")

    print("Switchboard risk calculation running...")        

    return {"message": f"Updated {updated_count} switchboard records successfully"}


    # return SwitchboardRiskResponse(
    #         risk_score=sb_data.risk_score,
    #         risk_level=sb_data.risk_level,
    #         created_at=sb_data.risk_created_at.isoformat()
    #     )


@app.get("/get_system_ip")
def get_system_ip():
    try:
        # Get local system IP
        hostname = socket.gethostname()
        local_ip = socket.gethostbyname(hostname)
        return {"ip_address": local_ip}
    except Exception as e:
        return {"error": str(e)}
    

#--------------------------------------------------------------------------------------------------
class LayoutSaveRequest(BaseModel):
    components: List[schemas.ComponentBase]


@app.post("/save_components", response_model=schemas.MessageResponse)
async def save_components(request: schemas.ComponentCreateRequest, db: Session = Depends(get_db)):
    user_name = request.user_name
    components = request.components

    # Ensure user exists
    user_exists = db.query(models.User).filter(models.User.email == user_name).first()
    if not user_exists:
        raise HTTPException(status_code=404, detail="User not found")

    if not components:
        raise HTTPException(status_code=400, detail="No components provided")

    # Identify unique floors
    floor_names = {comp.floor_name for comp in components}

    for floor_name in floor_names:
        # Fetch existing components for this user and floor
        existing_components = db.query(models.ComponentLayout).filter(
            models.ComponentLayout.user_name == user_name,
            models.ComponentLayout.floor_name == floor_name
        ).all()

        # Build key sets for comparison
        existing_keys = {(comp.floor_name, comp.instance_id) for comp in existing_components}
        new_keys = {(comp.floor_name, comp.instance_id) for comp in components if comp.floor_name == floor_name}

        # DELETE missing components from relevant tables
        for comp in existing_components:
            if (comp.floor_name, comp.instance_id) not in new_keys:
                component_type = comp.component_name.lower()

                # Delete from ComponentLayout
                db.delete(comp)

                if component_type == "wiring":
                    item = db.query(models.WiringEquipmentData).filter(
                        models.WiringEquipmentData.email == user_name,
                        models.WiringEquipmentData.floor_name == comp.floor_name,
                        models.WiringEquipmentData.instance_id == comp.instance_id
                    ).first()
                    if item:
                        db.delete(item)

                elif component_type == "panel":
                    item = db.query(models.PanelRiskData).filter(
                        models.PanelRiskData.email == user_name,
                        models.PanelRiskData.floor_name == comp.floor_name,
                        models.PanelRiskData.instance_id == comp.instance_id
                    ).first()
                    if item:
                        db.delete(item)

                elif component_type == "ac":
                    item = db.query(models.ACData).filter(
                        models.ACData.email == user_name,
                        models.ACData.floor_name == comp.floor_name,
                        models.ACData.instance_id == comp.instance_id
                    ).first()
                    if item:
                        db.delete(item)

                elif component_type == "battery":
                    item = db.query(models.BatteryLiveData).filter(
                        models.BatteryLiveData.email == user_name,
                        models.BatteryLiveData.floor_name == comp.floor_name,
                        models.BatteryLiveData.instance_id == comp.instance_id
                        

                    ).first()
                    if item:
                        db.delete(item)

                elif component_type == "ups":
                    item = db.query(models.UPSData).filter(
                        models.UPSData.email == user_name,
                        models.UPSData.floor_name == comp.floor_name,
                        models.UPSData.instance_id == comp.instance_id
                    ).first()
                    if item:
                        db.delete(item)

                elif component_type == "switchboard":
                    item = db.query(models.SwitchboardLiveData).filter(
                        models.SwitchboardLiveData.email == user_name,
                        models.SwitchboardLiveData.floor_name == comp.floor_name,
                        models.SwitchboardLiveData.instance_id == comp.instance_id
                    ).first()
                    if item:
                        db.delete(item)

        # # INSERT or UPDATE per floor
        # for comp in components:
        #     if comp.floor_name != floor_name:
        #         continue


        # INSERT or UPDATE per floor
        for comp in components:
            if comp.floor_name != floor_name:
                continue

            #Skip invalid/empty components to avoid blank entries
            if not comp.component_name or not comp.instance_id:
                continue
    

            # ---- ComponentLayout ----
            existing_component = db.query(models.ComponentLayout).filter(
                models.ComponentLayout.user_name == user_name,
                models.ComponentLayout.floor_name == comp.floor_name,
                models.ComponentLayout.instance_id == comp.instance_id
            ).first()

            if existing_component:
                existing_component.position_x = comp.position_x
                existing_component.position_y = comp.position_y
                existing_component.component_name = comp.component_name
                existing_component.grid_number = comp.grid_number
                # existing_component.location = comp.location


            else:
                new_component = models.ComponentLayout(
                    user_name=user_name,
                    floor_name=comp.floor_name,
                    component_name=comp.component_name,
                    instance_id=comp.instance_id,
                    position_x=comp.position_x,
                    position_y=comp.position_y,
                    grid_number=comp.grid_number,  
                    location=comp.location
                   
                    
                )
                db.add(new_component)

            # ----------Conditional Equipment Sync----------
            component_type = comp.component_name.lower()

            if component_type == "wiring":
                existing = db.query(models.WiringEquipmentData).filter(
                    models.WiringEquipmentData.email == user_name,
                    models.WiringEquipmentData.floor_name == comp.floor_name,
                    models.WiringEquipmentData.instance_id == comp.instance_id
                ).first()
                if not existing:
                    db.add(models.WiringEquipmentData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif component_type == "panel":
                existing = db.query(models.PanelRiskData).filter(
                    models.PanelRiskData.email == user_name,
                    models.PanelRiskData.floor_name == comp.floor_name,
                    models.PanelRiskData.instance_id == comp.instance_id
                ).first()
                if not existing:
                    db.add(models.PanelRiskData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif component_type == "ac":
                existing = db.query(models.ACData).filter(
                    models.ACData.email == user_name,
                    models.ACData.floor_name == comp.floor_name,
                    models.ACData.instance_id == comp.instance_id
                ).first()
                if not existing:
                    db.add(models.ACData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id,created_at=datetime.now()
                    ))

            elif component_type == "battery":
                existing = db.query(models.BatteryLiveData).filter(
                    models.BatteryLiveData.email == user_name,
                    models.BatteryLiveData.floor_name == comp.floor_name,
                    models.BatteryLiveData.instance_id == comp.instance_id
                  
                ).first()
                if not existing:
                    db.add(models.BatteryLiveData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id,created_at=datetime.now()
                    ))

            elif component_type == "ups":
                existing = db.query(models.UPSData).filter(
                    models.UPSData.email == user_name,
                    models.UPSData.floor_name == comp.floor_name,
                    models.UPSData.instance_id == comp.instance_id
                ).first()
                if not existing:
                    db.add(models.UPSData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif component_type == "switchboard":
                existing = db.query(models.SwitchboardLiveData).filter(
                    models.SwitchboardLiveData.email == user_name,
                    models.SwitchboardLiveData.floor_name == comp.floor_name,
                    models.SwitchboardLiveData.instance_id == comp.instance_id
                ).first()
                if not existing:
                    db.add(models.SwitchboardLiveData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

    db.commit()
    return {"message": "Components and all related equipment data synced successfully (wiring/panel/ac/battery/ups/switchboard per floor)"}

# -------------------------------------------------------------------------------------------------------------------------------------------
# @app.post("/save_components", response_model=schemas.MessageResponse)
# async def save_components(request: schemas.ComponentCreateRequest, db: Session = Depends(get_db)):
#     user_name = request.user_name
#     components = request.components

#     # Ensure user exists
#     user_exists = db.query(models.User).filter(models.User.email == user_name).first()
#     if not user_exists:
#         raise HTTPException(status_code=404, detail="User not found")

#     if not components:
#         raise HTTPException(status_code=400, detail="No components provided")

#     # Identify unique floors
#     floor_names = {comp.floor_name for comp in components}

#     for floor_name in floor_names:
#         # Fetch all existing components for this user and floor
#         existing_components = db.query(models.ComponentLayout).filter(
#             models.ComponentLayout.user_name == user_name,
#             models.ComponentLayout.floor_name == floor_name
#         ).all()

#         # --- Group existing and new components by component type ---
#         existing_by_type = defaultdict(set)
#         for comp in existing_components:
#             existing_by_type[comp.component_name.lower()].add(comp.instance_id)

#         new_by_type = defaultdict(set)
#         for comp in components:
#             if comp.floor_name == floor_name:
#                 new_by_type[comp.component_name.lower()].add(comp.instance_id)

#         # --- DELETE missing components (only same component type) ---
#         for component_type, existing_ids in existing_by_type.items():
#             new_ids = new_by_type.get(component_type, set())

#             for comp in existing_components:
#                 if comp.component_name.lower() == component_type and comp.instance_id not in new_ids:
#                     db.delete(comp)

#                     # Delete from corresponding equipment tables
#                     if component_type == "wiring":
#                         item = db.query(models.WiringEquipmentData).filter(
#                             models.WiringEquipmentData.email == user_name,
#                             models.WiringEquipmentData.floor_name == comp.floor_name,
#                             models.WiringEquipmentData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#                     elif component_type == "panel":
#                         item = db.query(models.PanelRiskData).filter(
#                             models.PanelRiskData.email == user_name,
#                             models.PanelRiskData.floor_name == comp.floor_name,
#                             models.PanelRiskData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#                     elif component_type == "ac":
#                         item = db.query(models.ACData).filter(
#                             models.ACData.email == user_name,
#                             models.ACData.floor_name == comp.floor_name,
#                             models.ACData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#                     elif component_type == "battery":
#                         item = db.query(models.BatteryLiveData).filter(
#                             models.BatteryLiveData.email == user_name,
#                             models.BatteryLiveData.floor_name == comp.floor_name,
#                             models.BatteryLiveData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#                     elif component_type == "ups":
#                         item = db.query(models.UPSData).filter(
#                             models.UPSData.email == user_name,
#                             models.UPSData.floor_name == comp.floor_name,
#                             models.UPSData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#                     elif component_type == "switchboard":
#                         item = db.query(models.SwitchboardLiveData).filter(
#                             models.SwitchboardLiveData.email == user_name,
#                             models.SwitchboardLiveData.floor_name == comp.floor_name,
#                             models.SwitchboardLiveData.instance_id == comp.instance_id
#                         ).first()
#                         if item:
#                             db.delete(item)

#         # --- INSERT or UPDATE components for this floor ---
#         for comp in components:
#             if comp.floor_name != floor_name:
#                 continue

#             # Skip invalid/empty entries
#             if not comp.component_name or not comp.instance_id:
#                 continue

#             # Update existing component or add new
#             existing_component = db.query(models.ComponentLayout).filter(
#                 models.ComponentLayout.user_name == user_name,
#                 models.ComponentLayout.floor_name == comp.floor_name,
#                 models.ComponentLayout.instance_id == comp.instance_id
#             ).first()

#             if existing_component:
#                 existing_component.position_x = comp.position_x
#                 existing_component.position_y = comp.position_y
#                 existing_component.component_name = comp.component_name
#                 existing_component.grid_number = comp.grid_number
#             else:
#                 new_component = models.ComponentLayout(
#                     user_name=user_name,
#                     floor_name=comp.floor_name,
#                     component_name=comp.component_name,
#                     instance_id=comp.instance_id,
#                     position_x=comp.position_x,
#                     position_y=comp.position_y,
#                     grid_number=comp.grid_number
#                 )
#                 db.add(new_component)

#             # --- Sync corresponding equipment data table ---
#             component_type = comp.component_name.lower()

#             if component_type == "wiring":
#                 existing = db.query(models.WiringEquipmentData).filter(
#                     models.WiringEquipmentData.email == user_name,
#                     models.WiringEquipmentData.floor_name == comp.floor_name,
#                     models.WiringEquipmentData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.WiringEquipmentData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#             elif component_type == "panel":
#                 existing = db.query(models.PanelRiskData).filter(
#                     models.PanelRiskData.email == user_name,
#                     models.PanelRiskData.floor_name == comp.floor_name,
#                     models.PanelRiskData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.PanelRiskData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#             elif component_type == "ac":
#                 existing = db.query(models.ACData).filter(
#                     models.ACData.email == user_name,
#                     models.ACData.floor_name == comp.floor_name,
#                     models.ACData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.ACData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#             elif component_type == "battery":
#                 existing = db.query(models.BatteryLiveData).filter(
#                     models.BatteryLiveData.email == user_name,
#                     models.BatteryLiveData.floor_name == comp.floor_name,
#                     models.BatteryLiveData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.BatteryLiveData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#             elif component_type == "ups":
#                 existing = db.query(models.UPSData).filter(
#                     models.UPSData.email == user_name,
#                     models.UPSData.floor_name == comp.floor_name,
#                     models.UPSData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.UPSData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#             elif component_type == "switchboard":
#                 existing = db.query(models.SwitchboardLiveData).filter(
#                     models.SwitchboardLiveData.email == user_name,
#                     models.SwitchboardLiveData.floor_name == comp.floor_name,
#                     models.SwitchboardLiveData.instance_id == comp.instance_id
#                 ).first()
#                 if not existing:
#                     db.add(models.SwitchboardLiveData(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ))

#     db.commit()
#     return {"message": "Components and related equipment data synced successfully (per type per floor)"}

#--------------------------------------------------------------------------------------------
class GetComponentsRequest(BaseModel):
    user_name: str

@app.post("/get_components")
async def get_components(request: GetComponentsRequest, db: Session = Depends(get_db)):
    user_name = request.user_name

    # Check if user exists
    user_exists = db.query(models.User).filter(models.User.email == user_name).first()
    if not user_exists:
        raise HTTPException(status_code=404, detail=f"User '{user_name}' not found")

    # Fetch all components for this user
    components = db.query(models.ComponentLayout).filter(models.ComponentLayout.user_name == user_name).all()

    if not components:
        raise HTTPException(status_code=404, detail="No components found for this user")

    return components

#Request schema
class UserEmailRequest(BaseModel):
    email: EmailStr

#Panel response
class PanelDetails(BaseModel):
    instance_id: str
    # risk_score: float
    # risk_level: str
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    floor_name: str
    email: str

#UPS response
class UPSDetails(BaseModel):
    instance_id: str
    # risk_score: float
    # risk_level: str
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    floor_name: str
    risk_created_at: str
    created_at: datetime
    email: str
   

class ACDetails(BaseModel):
    instance_id: str
    # risk_score: float
    # risk_level: str
    floor_name: str
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    created_at: datetime
    email: str

class BatteryDetails(BaseModel):
    instance_id: str
    # risk_score: float
    # risk_level: str  
    floor_name: str
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    created_at: datetime
    email: str

class SwitchboardDetails(BaseModel):
    instance_id: str
    # risk_score: float
    # risk_level: str
    floor_name: str
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    email: str

class WiringEquipmentDetails(BaseModel):
    instance_id: str
    floor_name: str
    risk_score: float
    risk_score: Optional[float] = None
    risk_level: Optional[str] = None
    # risk_level: str    
    email: str

#Combined response
class UserPanelUPSResponse(BaseModel):
    token_number: Optional[str] = None  
    panels: List[PanelDetails] = []
    ups_data: List[UPSDetails] = []
    ac_data: List[ACDetails] = []
    battery_data: List[BatteryDetails] = []
    switchboard_data: List[SwitchboardDetails]  
    wiring_equipment_data: List[WiringEquipmentDetails] = []  

# #API to fetch from 5 tables: panel, ups, ac, battery, switchboard
# @app.post("/user_panel_details", response_model=UserPanelUPSResponse)
# def get_user_panel_details(request: UserEmailRequest, db: Session = Depends(get_db)):
#     # --- Fetch from panel_risk_data ---
#     panel_records = (
#         db.query(
#             PanelRiskData.instance_id,
#             PanelRiskData.risk_score,
#             PanelRiskData.risk_level,
#             PanelRiskData.floor_name
#         )
#         .filter(PanelRiskData.email == request.email)
#         .all()
#     )

#     # --- Fetch from ups_data ---
#     ups_records = (
#         db.query(
#             UPSData.instance_id,
#             UPSData.risk_score,
#             UPSData.risk_level,
#             UPSData.risk_created_at,
#             UPSData.floor_name
#         )
#         .filter(UPSData.email == request.email)
#         .all()
#     )

#     # --- Fetch from ac_data ---
#     ac_records = (
#         db.query(
#             ACData.instance_id,
#             ACData.risk_score,
#             ACData.risk_level,
#             ACData.floor_name
#         )
#         .filter(ACData.email == request.email)
#         .all()
#     )

#     # --- Fetch from battery_live_data ---
#     battery_records = (
#         db.query(
#             BatteryLiveData.instance_id,
#             BatteryLiveData.risk_score,
#             BatteryLiveData.risk_level,
#             BatteryLiveData.floor_name,
#             BatteryLiveData.created_at
#         )
#         .filter(BatteryLiveData.email == request.email)
#         .all()
#     )

#     # --- Fetch from switchboard_live_data ---
#     switchboard_records = (
#         db.query(
#             SwitchboardLiveData.instance_id,
#             SwitchboardLiveData.risk_score,
#             SwitchboardLiveData.risk_level,
#             SwitchboardLiveData.floor_name
#         )
#         .filter(SwitchboardLiveData.email == request.email)
#         .all()
#     )

#     # --- Fetch from wiring_equipment_data ---
#     wiring_equipment_records = (
#         db.query(
#             WiringEquipmentData.instance_id,
#             WiringEquipmentData.floor_name, 
#             WiringEquipmentData.risk_score,
#             WiringEquipmentData.risk_level
#         )
#         .filter(WiringEquipmentData.email == request.email)
#         .all()
#     )

#     # --- If no records found in any table ---
#     if not panel_records and not ups_records and not ac_records and not battery_records and not switchboard_records and not wiring_equipment_records:
#         raise HTTPException(status_code=404, detail="No records found for this email")

#     return UserPanelUPSResponse(
#         panels=[
#             PanelDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#             )
#             for r in panel_records
#         ],
#         ups_data=[
#             UPSDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 risk_created_at=str(r.risk_created_at),
#                 floor_name=r.floor_name,
#             )
#             for r in ups_records
#         ],
#         ac_data=[
#             ACDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#             )
#             for r in ac_records
#         ],
#         battery_data=[
#             BatteryDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#                 created_at=r.created_at
#             )
#             for r in battery_records
#         ],
#         switchboard_data=[
#             SwitchboardDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#             )
#             for r in switchboard_records
#         ],
#         wiring_equipment_data=[
#             WiringEquipmentDetails(
#                 instance_id=r.instance_id,
#                 floor_name=r.floor_name,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level
#             )
#             for r in wiring_equipment_records
#         ]
#     )



#----------------------user_panel_details_working_code---------------------------------
# #API to fetch from 5 tables: panel, ups, ac, battery, switchboard
# @app.post("/user_panel_details", response_model=UserPanelUPSResponse)
# def get_user_panel_details(request: UserEmailRequest, db: Session = Depends(get_db)):
#     # --- Fetch from panel_risk_data ---
#     panel_records = (
#         db.query(
#             PanelRiskData.instance_id,
#             PanelRiskData.risk_score,
#             PanelRiskData.risk_level,
#             PanelRiskData.floor_name,
#             PanelRiskData.email
#         )
#         .filter(PanelRiskData.email == request.email)
#         .all()
#     )

#     # --- Fetch from ups_data ---
#     ups_records = (
#         db.query(
#             UPSData.instance_id,
#             UPSData.risk_score,
#             UPSData.risk_level,
#             UPSData.risk_created_at,
#             UPSData.floor_name,
#             UPSData.created_at,
#             UPSData.email
           
#         )
#         .filter(UPSData.email == request.email)
#         .all()
#     )

#     # --- Fetch from ac_data ---
#     ac_records = (
#         db.query(
#             ACData.instance_id,
#             ACData.risk_score,
#             ACData.risk_level,
#             ACData.floor_name,
#             ACData.created_at,
#             ACData.email
#         )
#         .filter(ACData.email == request.email)
#         .all()
#     )

#     # # Subquery: Get latest created_at per instance_id & floor_name
#     # subquery = (
#     #     db.query(
#     #         BatteryLiveData.instance_id,
#     #         BatteryLiveData.floor_name,
#     #         func.max(BatteryLiveData.created_at).label("latest_created_at")
#     #     )
#     #     .filter(BatteryLiveData.email == request.email)
#     #     .group_by(BatteryLiveData.instance_id, BatteryLiveData.floor_name)
#     #     .subquery()
#     # )

#     # Join back to get the full record for those latest entries
#     battery_records = (
#         db.query(
#             BatteryLiveData.instance_id,
#             BatteryLiveData.floor_name,
#             BatteryLiveData.risk_score,
#             BatteryLiveData.risk_level,
#             BatteryLiveData.created_at,
#             BatteryLiveData.email
#         )
#         # .join(
#         #     subquery,
#         #     (BatteryLiveData.instance_id == subquery.c.instance_id)
#         #     & (BatteryLiveData.floor_name == subquery.c.floor_name)
#         #     & (BatteryLiveData.created_at == subquery.c.latest_created_at)
#         # )
#         # .filter(BatteryLiveData.email == request.email)
#         # .order_by(desc(BatteryLiveData.created_at))
#         # .all()

#         .filter(BatteryLiveData.email == request.email)
#         .all()
#     )

#     # --- Fetch from switchboard_live_data ---
#     switchboard_records = (
#         db.query(
#             SwitchboardLiveData.instance_id,
#             SwitchboardLiveData.risk_score,
#             SwitchboardLiveData.risk_level,
#             SwitchboardLiveData.floor_name,
#             SwitchboardLiveData.email
#         )
#         .filter(SwitchboardLiveData.email == request.email)
#         .all()
#     )

#     # --- Fetch from wiring_equipment_data ---
#     wiring_equipment_records = (
#         db.query(
#             WiringEquipmentData.instance_id,
#             WiringEquipmentData.floor_name, 
#             WiringEquipmentData.risk_score,
#             WiringEquipmentData.risk_level,
#             WiringEquipmentData.email
#         )
#         .filter(WiringEquipmentData.email == request.email)
#         .all()
#     )

#     # --- If no records found in any table ---
#     if not panel_records and not ups_records and not ac_records and not battery_records and not switchboard_records and not wiring_equipment_records:
#         raise HTTPException(status_code=404, detail="No records found for this email")

#     return UserPanelUPSResponse(
#         panels=[
#             PanelDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#                 email=r.email
#             )
#             for r in panel_records
#         ],
#         ups_data=[
#             UPSDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 risk_created_at=str(r.risk_created_at),
#                 floor_name=r.floor_name,
#                 created_at=r.created_at,
#                 email=r.email   
#             )
#             for r in ups_records
#         ],
#         ac_data=[
#             ACDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#                 created_at=r.created_at,
#                 email=r.email
#             )
#             for r in ac_records
#         ],
#         battery_data=[
#             BatteryDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#                 created_at=r.created_at,
#                 email=r.email
#             )
#             for r in battery_records
#         ],
#         switchboard_data=[
#             SwitchboardDetails(
#                 instance_id=r.instance_id,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 floor_name=r.floor_name,
#                 email=r.email
#             )
#             for r in switchboard_records
#         ],
#         wiring_equipment_data=[
#             WiringEquipmentDetails(
#                 instance_id=r.instance_id,
#                 floor_name=r.floor_name,
#                 risk_score=r.risk_score,
#                 risk_level=r.risk_level,
#                 email=r.email
#             )
#             for r in wiring_equipment_records
#         ]
#     )




@app.post("/user_panel_details", response_model=UserPanelUPSResponse)
def get_user_panel_details(request: UserEmailRequest, db: Session = Depends(get_db)):

    # ------------------------------
    # 1. GET TOKEN FROM USERS TABLE
    # ------------------------------
    user_data = db.query(models.User).filter(models.User.email == request.email).first()
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")

    token_number = user_data.token_number

    # ------------------------------
    # 2. FETCH PANEL DATA
    # ------------------------------
    panel_records = (
        db.query(
            PanelRiskData.instance_id,
            PanelRiskData.risk_score,
            PanelRiskData.risk_level,
            PanelRiskData.floor_name,
            PanelRiskData.email
        )
        .filter(PanelRiskData.email == request.email)
        .all()
    )

    # ------------------------------
    # 3. FETCH UPS DATA
    # ------------------------------
    ups_records = (
        db.query(
            UPSData.instance_id,
            UPSData.risk_score,
            UPSData.risk_level,
            UPSData.risk_created_at,
            UPSData.floor_name,
            UPSData.created_at,
            UPSData.email
        )
        .filter(UPSData.email == request.email)
        .all()
    )

    # ------------------------------
    # 4. FETCH AC DATA
    # ------------------------------
    ac_records = (
        db.query(
            ACData.instance_id,
            ACData.risk_score,
            ACData.risk_level,
            ACData.floor_name,
            ACData.created_at,
            ACData.email
        )
        .filter(ACData.email == request.email)
        .all()
    )

    # ------------------------------
    # 5. FETCH BATTERY DATA
    # ------------------------------
    battery_records = (
        db.query(
            BatteryLiveData.instance_id,
            BatteryLiveData.floor_name,
            BatteryLiveData.risk_score,
            BatteryLiveData.risk_level,
            BatteryLiveData.created_at,
            BatteryLiveData.email
        )
        .filter(BatteryLiveData.email == request.email)
        .all()
    )

    # ------------------------------
    # 6. FETCH SWITCHBOARD DATA
    # ------------------------------
    switchboard_records = (
        db.query(
            SwitchboardLiveData.instance_id,
            SwitchboardLiveData.risk_score,
            SwitchboardLiveData.risk_level,
            SwitchboardLiveData.floor_name,
            SwitchboardLiveData.email
        )
        .filter(SwitchboardLiveData.email == request.email)
        .all()
    )

    # ------------------------------
    # 7. FETCH WIRING EQUIPMENT DATA
    # ------------------------------
    wiring_equipment_records = (
        db.query(
            WiringEquipmentData.instance_id,
            WiringEquipmentData.floor_name,
            WiringEquipmentData.risk_score,
            WiringEquipmentData.risk_level,
            WiringEquipmentData.email
        )
        .filter(WiringEquipmentData.email == request.email)
        .all()
    )

    # ------------------------------
    # 8. CHECK EMPTY RESULTS
    # ------------------------------
    if not (panel_records or ups_records or ac_records or battery_records or switchboard_records or wiring_equipment_records):
        raise HTTPException(status_code=404, detail="No records found for this email")

    # ------------------------------
    # 9. RETURN COMPLETE RESPONSE
    # ------------------------------
    return UserPanelUPSResponse(
        token_number=token_number,

        panels=[
            PanelDetails(
                instance_id=r.instance_id,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                floor_name=r.floor_name,
                email=r.email
            ) for r in panel_records
        ],

        ups_data=[
            UPSDetails(
                instance_id=r.instance_id,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                risk_created_at=str(r.risk_created_at),
                floor_name=r.floor_name,
                created_at=r.created_at,
                email=r.email
            ) for r in ups_records
        ],

        ac_data=[
            ACDetails(
                instance_id=r.instance_id,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                floor_name=r.floor_name,
                created_at=r.created_at,
                email=r.email
            ) for r in ac_records
        ],

        battery_data=[
            BatteryDetails(
                instance_id=r.instance_id,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                floor_name=r.floor_name,
                created_at=r.created_at,
                email=r.email
            ) for r in battery_records
        ],

        switchboard_data=[
            SwitchboardDetails(
                instance_id=r.instance_id,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                floor_name=r.floor_name,
                email=r.email
            ) for r in switchboard_records
        ],

        wiring_equipment_data=[
            WiringEquipmentDetails(
                instance_id=r.instance_id,
                floor_name=r.floor_name,
                risk_score=r.risk_score,
                risk_level=r.risk_level,
                email=r.email
            ) for r in wiring_equipment_records
        ]
    )







class RiskCalculationRequest(BaseModel):
    instance_id: str

def sigmoid(x):
    return 1 / (1 + exp(-x))

# @app.post("/calculate_wiring_risk")
# def calculate_wiring_risk(request: RiskCalculationRequest):
#     db: Session = next(get_db())
#     instance_id = request.instance_id

#     #Fetch latest sensor data for this panel
#     sensor_data = db.query(WiringEquipmentData).filter(WiringEquipmentData.instance_id == instance_id).first()
#     if not sensor_data:
#         raise HTTPException(status_code=404, detail="No sensor data found for the panel")

#     #Assign sensor values
#     I = sensor_data.current
#     Vdrop = sensor_data.voltage_drop
#     Ileak = sensor_data.leakage_current
#     ΔT = sensor_data.temperature_rise
#     Ifault = sensor_data.fault_current
#     I2t = sensor_data.i2t
#     Psurge = sensor_data.surge_power
#     EnvIdx = sensor_data.environment_index

#     #Fixed reference / normalization values (configured in code)
#     P_rated = 3000.0        # W
#     Vdrop_max = 3.0         # V
#     Ileak_ref = 0.005       # A
#     T_margin = 40.0         # °C
#     Ifault_ref = 0.03       # A
#     I2t_ref = 5000.0        # A²·s
#     Psurge_max = 200.0      # J
#     EnvIdx_max = 1.0        # dimensionless
#     R = 0.5                 # Ω

#     #Weights (optional)
#     w1 = w2 = w3 = w4 = w5 = w6 = w7 = w8 = w9 = 1.0

#     #Calculate individual terms
#     t1 = w1 * (I**2 * R / P_rated)
#     t2 = w2 * (Vdrop / Vdrop_max)
#     t3 = w3 * (Ileak / Ileak_ref)
#     t4 = 0.0  # Removed Life_ref/Life_current to avoid dominating the score
#     t5 = w5 * (ΔT / T_margin)
#     t6 = w6 * (Ifault / Ifault_ref)
#     t7 = w7 * (I2t / I2t_ref)
#     t8 = w8 * (Psurge / Psurge_max)
#     t9 = w9 * (EnvIdx / EnvIdx_max)

#     #Raw score
#     x = t1 + t2 + t3 + t4 + t5 + t6 + t7 + t8 + t9

#     #Sigmoid scaling to 0..1
#     risk_score = sigmoid(x)

#     #Determine risk level
#     if risk_score < 0.4:
#         risk_level = "Normal"
#     elif risk_score < 0.7:
#         risk_level = "Medium"
#     else:
#         risk_level = "High"

#     #Update the record with risk values
#     sensor_data.risk_score = risk_score
#     sensor_data.risk_level = risk_level
#     sensor_data.risk_created_at = datetime.utcnow()
#     db.commit()

#     #Return response
#     return {
#         "instance_id": instance_id,
#         "risk_score": risk_score,
#         "risk_level": risk_level,
#     }

# ---------------------Calculation of all AC risk working code -------------------------
# FARI calculation function (unchanged)
def calculate_fari(ac_data):
    α1, α2 = 0.7, 0.3
    β1, β2 = 0.6, 0.4
    W = [2.0, 1.5, 1.2, 1.0, 1.3]


    EAI = α1 * abs(ac_data.I_measured - ac_data.Irated) / ac_data.Irated + α2 * (ac_data.THD / ac_data.THDmax)
    TSI = (ac_data.Tc - ac_data.Tsafe) / ac_data.Tsafe
    CHI = β1 * (ac_data.ESR / ac_data.ESRlimit) + β2 * ac_data.C_drop
    MSI = ac_data.RHpcb / ac_data.RHthreshold
    CIRI = (ac_data.Ires / ac_data.Ilimit) + (1 / ac_data.IR)

    FARIraw = W[0]*EAI + W[1]*TSI + W[2]*CHI + W[3]*MSI + W[4]*CIRI
    FARInorm = FARIraw / sum(W)

    # if FARInorm < 0.5:
    #     risk_level = "Normal"
    # elif 0.5 <= FARInorm < 1.0:
    #     risk_level = "Medium"
    # elif 1.0 <= FARInorm < 1.5:
    #     risk_level = "High"
    # else:
    #     risk_level = "High"


    if FARInorm < 0.5:
        risk_level = "Normal"
    elif 0.5 <= FARInorm < 1.0:
        risk_level = "Medium"
    else:
        risk_level = "High"
        
        
    
    # FARInorm = round(FARInorm * 100, 2)
    FARInorm = round((FARInorm / 1.5) * 100, 2)
    FARInorm = min(FARInorm, 100)

    return FARInorm, risk_level

# ----------------------------------------------------------------------------------------

# def calculate_fari(ac_data):
#     α1, α2 = 0.7, 0.3
#     β1, β2 = 0.6, 0.4
#     W = [2.0, 1.5, 1.2, 1.0, 1.3]

#     # Avoid division by zero safely
#     Irated = ac_data.Irated if ac_data.Irated and ac_data.Irated != 0 else 1e-6
#     THDmax = ac_data.THDmax if ac_data.THDmax and ac_data.THDmax != 0 else 1e-6
#     Tsafe = ac_data.Tsafe if ac_data.Tsafe and ac_data.Tsafe != 0 else 1e-6
#     ESRlimit = ac_data.ESRlimit if ac_data.ESRlimit and ac_data.ESRlimit != 0 else 1e-6
#     RHthreshold = ac_data.RHthreshold if ac_data.RHthreshold and ac_data.RHthreshold != 0 else 1e-6
#     Ilimit = ac_data.Ilimit if ac_data.Ilimit and ac_data.Ilimit != 0 else 1e-6
#     IR = ac_data.IR if ac_data.IR and ac_data.IR != 0 else 1e-6

#     # Calculate indexes
#     EAI = α1 * abs(ac_data.I_measured - Irated) / Irated + α2 * (ac_data.THD / THDmax)
#     TSI = (ac_data.Tc - ac_data.Tsafe) / Tsafe
#     CHI = β1 * (ac_data.ESR / ESRlimit) + β2 * ac_data.C_drop
#     MSI = ac_data.RHpcb / RHthreshold
#     CIRI = (ac_data.Ires / Ilimit) + (1 / IR)

#     # Combine all weighted indices
#     FARIraw = W[0]*EAI + W[1]*TSI + W[2]*CHI + W[3]*MSI + W[4]*CIRI
#     FARInorm = FARIraw / sum(W)

#     # Classify risk level
#     if FARInorm < 0.5:
#         risk_level = "Normal"
#     elif 0.5 <= FARInorm < 1.0:
#         risk_level = "Medium"
#     else:
#         risk_level = "High"

#     # Normalize percentage to 0–100
#     FARInorm = round((FARInorm / 1.5) * 100, 2)
#     FARInorm = min(FARInorm, 100)

#     return FARInorm, risk_level

#Updated API endpoint
@app.post("/calculate_all_ac_risks")
def calculate_all_ac_risks(db: Session = Depends(get_db)):
    # Fetch all AC instances
    all_ac_data = db.query(ACData).all()

    if not all_ac_data:
        raise HTTPException(status_code=404, detail="No AC data found")

    results = []

    # Calculate and update each record
    for ac_data in all_ac_data:
        fari_score, risk_level = calculate_fari(ac_data)

        ac_data.risk_score = round(fari_score, 3)
        ac_data.risk_level = risk_level
        ac_data.risk_created_at = datetime.utcnow()

        results.append({
            "instance_id": ac_data.instance_id,
            "risk_score": ac_data.risk_score,
            "risk_level": ac_data.risk_level,
            "updated_at": ac_data.risk_created_at.isoformat()
        })

    #Commit all updates once
    db.commit()

    print("AC risk calculation running...")
    return {"status": "success", "updated_count": len(results), "results": results}


@app.post("/wiring_loop_calculation")
def calculate_wiring_risk_all():
    db: Session = next(get_db())

    # Fetch all sensor data rows
    all_sensor_data = db.query(WiringEquipmentData).all()
    if not all_sensor_data:
        raise HTTPException(status_code=404, detail="No sensor data found")

    results = []

    # Fixed reference / normalization values
    P_rated = 3000.0
    Vdrop_max = 3.0
    Ileak_ref = 0.005
    T_margin = 40.0
    Ifault_ref = 0.03
    I2t_ref = 5000.0
    Psurge_max = 200.0
    EnvIdx_max = 1.0
    R = 0.5

    # Weights
    w1 = w2 = w3 = w4 = w5 = w6 = w7 = w8 = w9 = 1.0

    # Function to calculate sigmoid
    def sigmoid(x):
        import math
        return 1 / (1 + math.exp(-x))

    for sensor_data in all_sensor_data:
        
        # Assign sensor values
        I = sensor_data.current
        Vdrop = sensor_data.voltage_drop
        Ileak = sensor_data.leakage_current
        ΔT = sensor_data.temperature_rise
        Ifault = sensor_data.fault_current
        I2t = sensor_data.i2t
        Psurge = sensor_data.surge_power
        EnvIdx = sensor_data.environment_index

        # Calculate terms
        t1 = w1 * (I**2 * R / P_rated)
        t2 = w2 * (Vdrop / Vdrop_max)
        t3 = w3 * (Ileak / Ileak_ref)
        t4 = 0.0
        t5 = w5 * (ΔT / T_margin)
        t6 = w6 * (Ifault / Ifault_ref)
        t7 = w7 * (I2t / I2t_ref)
        t8 = w8 * (Psurge / Psurge_max)
        t9 = w9 * (EnvIdx / EnvIdx_max)

        x = t1 + t2 + t3 + t4 + t5 + t6 + t7 + t8 + t9
        risk_score = sigmoid(x)

        # Determine risk level
        if risk_score < 0.4:
            risk_level = "Normal"
        elif risk_score < 0.7:
            risk_level = "Medium"
        else:
            risk_level = "High"

        # Update record
        sensor_data.risk_score = risk_score
        sensor_data.risk_level = risk_level
        sensor_data.risk_created_at = datetime.utcnow()

        results.append({
            "instance_id": sensor_data.instance_id,
            "floor_name": sensor_data.floor_name, 
            "risk_score": risk_score,
            "risk_level": risk_level
        })

    # Commit all updates at once
    db.commit()
    print("Wiring loop calculation running...")
    return {"updated_panels": results, "total": len(results)}


# Request schema
class FloorDataCreate(BaseModel):
    name: str
    email: EmailStr
    num_floors: int


# API to add or update floor data
# @app.post("/add_floor_data")
# def add_or_update_floor_data(request: FloorDataCreate, db: Session = Depends(get_db)):
#     # Check if the email already exists
#     existing = db.query(models.FloorData).filter(models.FloorData.email == request.email).first()

#     if existing:
#         #Update existing record
#         existing.name = request.name
#         existing.num_floors = request.num_floors
#         db.commit()
#         db.refresh(existing)
#         return {"message": f"Updated: {request.num_floors} floors saved successfully for {request.email}"}
#     else:
#         # Create new record
#         new_data = models.FloorData(
#             name=request.name,
#             email=request.email,
#             num_floors=request.num_floors
#         )
#         db.add(new_data)
#         db.commit()
#         return {"message": f"New record created: {request.num_floors} floors saved successfully for {request.email}"}




# @app.post("/add_floor_data")
# def add_or_update_floor_data(request: FloorDataCreate, db: Session = Depends(get_db)):
#     #  Check if user exists
#     user_exists = db.query(models.User).filter(models.User.email == request.email).first()
#     if not user_exists:
#         raise HTTPException(status_code=404, detail="User not found")

#     #  Check if floor data already exists for this email
#     existing = db.query(models.FloorData).filter(models.FloorData.email == request.email).first()

#     if existing:
#         # Update existing record
#         existing.name = request.name
#         existing.num_floors = request.num_floors
#         db.commit()
#         db.refresh(existing)
#         return {"message": f"Updated: {request.num_floors} floors saved successfully for {request.email}"}
#     else:
#         # Create new record
#         new_data = models.FloorData(
#             name=request.name,
#             email=request.email,
#             num_floors=request.num_floors
#         )
#         db.add(new_data)
#         db.commit()
#         return {"message": f"New record created: {request.num_floors} floors saved successfully for {request.email}"}

    
@app.post("/add_floor_data")
def add_or_update_floor_data(request: FloorDataCreate, db: Session = Depends(get_db)):
    #  Check if user exists
    user_exists = db.query(models.User).filter(models.User.email == request.email).first()
    if not user_exists:
        raise HTTPException(status_code=404, detail="User not found")

    #  Check if the user role is 'admin'
    if user_exists.role.lower() != "admin":
        raise HTTPException(status_code=403, detail="Access denied: only admin can add or update floor data")

    #  Check if floor data already exists for this email
    existing = db.query(models.FloorData).filter(models.FloorData.email == request.email).first()

    if existing:
        # Update existing record
        existing.name = request.name
        existing.num_floors = request.num_floors
        db.commit()
        db.refresh(existing)
        return {"message": f"Updated: {request.num_floors} floors saved successfully for {request.email}"}
    else:
        # Create new record
        new_data = models.FloorData(
            name=request.name,
            email=request.email,
            num_floors=request.num_floors
        )
        db.add(new_data)
        db.commit()
        return {"message": f"New record created: {request.num_floors} floors saved successfully for {request.email}"}




    
class EmailRequest(BaseModel):
    email: EmailStr

#API to fetch number of floors by email
@app.post("/get_floors_by_email")
def get_floors_by_email(request: EmailRequest, db: Session = Depends(get_db)):
    record = db.query(models.FloorData).filter(models.FloorData.email == request.email).first()
    if not record:
        raise HTTPException(status_code=404, detail="Email not found")

    return {"num_floors": record.num_floors}


# Response model
class AdminUser(BaseModel):
    name: str
    email: str
    company_name: str
    

@app.get("/admins", response_model=List[AdminUser])
def get_admin_list():
    try:
        with engine.connect() as conn:
            result = conn.execute(
                text("SELECT name, email, company_name  FROM users WHERE role = 'admin'")
            )
            admins = [dict(row._mapping) for row in result]
            if not admins:
                raise HTTPException(status_code=404, detail="No admins found")
            return admins
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




API_KEY = os.getenv("OPENWEATHER_API_KEY")
BASE_URL = "https://api.openweathermap.org/data/2.5/weather"

class WeatherRequest(BaseModel):
    location: str

@app.post("/weather_report_values")
def get_weather(request: WeatherRequest):
    location = request.location.strip()
    if not location:
        raise HTTPException(status_code=400, detail="Location cannot be empty")

    params = { 
        "q": location,
        "appid": API_KEY,
        "units": "metric"  
    }

    try:
        response = requests.get(BASE_URL, params=params)
        data = response.json()

        if response.status_code != 200:
            raise HTTPException(status_code=response.status_code, detail=data.get("message", "Error fetching weather data"))

        return {
            "location": f"{data['name']}, {data['sys']['country']}",
            "temperature": data["main"]["temp"],
            "weather": data["weather"][0]["description"],
            "humidity": data["main"]["humidity"],
            "wind_speed": data["wind"]["speed"]
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))    



# def vary(value: float, percent: float = 10.0):
#     """Return a slightly varied value (±percent%)."""
#     variation = value * percent / 100
#     return round(value + random.uniform(variation, variation), 2)

# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds with varied measured values."""
#     await asyncio.sleep(3)  # Wait for app startup

#     while True:
#         db = SessionLocal()
#         try:
#             # Get the latest battery record
#             latest = db.query(BatteryLiveData).order_by(BatteryLiveData.id.desc()).first()

#             if latest:
#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     # Varying only live measured/predicted values
#                     I_meas=vary(latest.I_meas),
#                     V_meas=vary(latest.V_meas),
#                     T_meas=vary(latest.T_meas),
#                     SOC_pred=vary(latest.SOC_pred),
#                     Cycle_pred=vary(latest.Cycle_pred),
#                     IR_pred=vary(latest.IR_pred),

#                     # Keep risk columns empty or same if you want
#                     risk_score=None,
#                     risk_level=None,
#                     risk_created_at=datetime.utcnow(),
#                 )

#                 db.add(new_record)
#                 db.commit()
#                 print(f"[{datetime.utcnow()}]  New battery data inserted for {latest.instance_id}")

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)  # Wait 5 seconds before next insert



def vary(value: float, percent: float = 15.0):
    """Return a slightly varied value (±percent%)."""
    variation = value * percent / 100
    # Use ±variation instead of same sign both sides
    return round(value + random.uniform(variation, variation), 2)

# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds with varied measured values,
#     and stop if risk_score > 1.000.
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     while True:
#         db: Session = SessionLocal()
#         try:
#             # Get the latest battery record
#             latest = db.query(BatteryLiveData).order_by(BatteryLiveData.id.desc()).first()

#             if not latest:
#                 print("No battery data found to base new records on.")
#                 break

#             # Stop condition: if risk_score exceeds 1.000
#             if latest.risk_score is not None and latest.risk_score > 100.0:
#                 print(f"Stopping updates: risk_score={latest.risk_score} exceeded 1.000")
#                 break

#             # Create a new record with slightly varied values
#             new_record = BatteryLiveData(
#                 instance_id=latest.instance_id,
#                 floor_name=latest.floor_name,
#                 email=latest.email,

#                 I_meas=vary(latest.I_meas),
#                 V_meas=vary(latest.V_meas),
#                 T_meas=vary(latest.T_meas),
#                 SOC_pred=vary(latest.SOC_pred),
#                 Cycle_pred=vary(latest.Cycle_pred),
#                 IR_pred=vary(latest.IR_pred),

#             )

#             db.add(new_record)
#             db.commit()
#             print(f"[{datetime.now()}]  New battery data inserted for {latest.instance_id}")

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)  # Wait 5 seconds before next insert


# from datetime import datetime

# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds with varied measured values,
#     and stop if risk_score > 1.000.
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     while True:
#         db: Session = SessionLocal()
#         try:
#             # Get the latest battery record
#             latest = db.query(BatteryLiveData).order_by(BatteryLiveData.id.desc()).first()

#             if not latest:
#                 print("No battery data found to base new records on.")
#                 break

#             # Stop condition: if risk_score exceeds 1.000
#             if latest.risk_score is not None and latest.risk_score > 100.0:
#                 print(f"Stopping updates: risk_score={latest.risk_score} exceeded 100.0")
#                 break

#             # Create a new record with slightly varied values
#             new_record = BatteryLiveData(
#                 instance_id=latest.instance_id,
#                 floor_name=latest.floor_name,
#                 email=latest.email,

#                 I_meas=vary(latest.I_meas),
#                 V_meas=vary(latest.V_meas),
#                 T_meas=vary(latest.T_meas),
#                 SOC_pred=vary(latest.SOC_pred),
#                 Cycle_pred=vary(latest.Cycle_pred),
#                 IR_pred=vary(latest.IR_pred),

#                 #  Store the current system time
#                 created_at=datetime.now()
#             )

#             db.add(new_record)
#             db.commit()
#             print(f"[{datetime.now()}]  New battery data inserted for {latest.instance_id}")

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)  # Wait 5 seconds before next insert

# -----------------------------working------------------------------------
# from datetime import datetime
# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds for all batteries.
#     Each battery stops updating individually when its risk_score > 100.0.
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     stopped_batteries = set()  # Keep track of batteries that crossed threshold

#     while True:
#         db: Session = SessionLocal()
#         try:
#             # Fetch all distinct instance IDs (batteries)
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]

#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 # Skip batteries that have already crossed the limit
#                 if instance_id in stopped_batteries:
#                     continue

#                 # Get latest record for each battery
#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )

#                 if not latest:
#                     continue

#                 # If this battery has crossed the threshold, skip it in future
#                 if latest.risk_score is not None and latest.risk_score > 100.0:
#                     print(f"[STOP] {instance_id} risk_score={latest.risk_score:.2f} exceeded 100.0")
#                     stopped_batteries.add(instance_id)
#                     continue

#                 # Create a new record with slightly varied values
#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     I_meas=vary(latest.I_meas),
#                     V_meas=vary(latest.V_meas),
#                     T_meas=vary(latest.T_meas),
#                     SOC_pred=vary(latest.SOC_pred),
#                     Cycle_pred=vary(latest.Cycle_pred),
#                     IR_pred=vary(latest.IR_pred),

#                     created_at=datetime.now(),
#                 )

#                 db.add(new_record)
#                 print(f"[{datetime.now()}] Inserted new data for {instance_id}")

#             db.commit()

#             # If all batteries have stopped, exit loop
#             if len(stopped_batteries) == len(instance_ids):
#                 print("All batteries crossed the threshold. Stopping updates.")
#                 break

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)





# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds for all batteries.
#     Each battery stops updating individually when its risk_score > 100.0.
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     stopped_batteries = set()  # Keep track of batteries that crossed threshold
#     day_offset = 0  # Counter to simulate days increasing

#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
#     # e.g., 2025-11-07 00:00:00

#     while True:
#         db: Session = SessionLocal()
#         try:
#             # Fetch all distinct instance IDs (batteries)
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]

#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 # Skip batteries that have already crossed the limit
#                 if instance_id in stopped_batteries:
#                     continue

#                 # Get latest record for each battery
#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )

#                 if not latest:
#                     continue

#                 # If this battery has crossed the threshold, skip it in future
#                 if latest.risk_score is not None and latest.risk_score > 100.0:
#                     print(f"[STOP] {instance_id} risk_score={latest.risk_score:.2f} exceeded 100.0")
#                     stopped_batteries.add(instance_id)
#                     continue

#                 # Simulated date: today's start date + offset days
#                 simulated_date = start_date + timedelta(days=day_offset)

#                 # Create a new record with slightly varied values
#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     I_meas=vary(latest.I_meas),
#                     V_meas=vary(latest.V_meas),
#                     T_meas=vary(latest.T_meas),
#                     SOC_pred=vary(latest.SOC_pred),
#                     Cycle_pred=vary(latest.Cycle_pred),
#                     IR_pred=vary(latest.IR_pred),

#                     created_at=simulated_date,
#                 )

#                 db.add(new_record)
#                 print(f"[{simulated_date}] Inserted new data for {instance_id}")

#             db.commit()

#             # After all inserts, increase simulated day count by 1
#             day_offset += 1

#             # If all batteries have stopped, exit loop
#             if len(stopped_batteries) == len(instance_ids):
#                 print("All batteries crossed the threshold. Stopping updates.")
#                 break

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)
# --------------------------------------------------------------------------------------------------


# async def auto_update_battery_data():
#     """Automatically insert new battery data every 5 seconds for all batteries.
#     Each battery stops updating individually when its risk_score > 100.0.
#     Handles initial zero values by applying a starting baseline increment.
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     stopped_batteries = set()  # Keep track of batteries that crossed threshold
#     day_offset = 0  # Counter to simulate days increasing

#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

#     while True:
#         db: Session = SessionLocal()
#         try:
#             # Fetch all distinct instance IDs (batteries)
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]

#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 if instance_id in stopped_batteries:
#                     continue

#                 # Get latest record for each battery
#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )

#                 if not latest:
#                     continue

#                 # Stop if above threshold
#                 if latest.risk_score is not None and latest.risk_score > 100.0:
#                     print(f"[STOP] {instance_id} risk_score={latest.risk_score:.2f} exceeded 100.0")
#                     stopped_batteries.add(instance_id)
#                     continue

#                 # Handle initial zero values (baseline increment)
#                 def baseline(value, base):
#                     return base if value == 0 else vary(value)

#                 simulated_date = start_date + timedelta(days=day_offset)

#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     # If first record has zeros, assign baseline start values
#                     I_meas=baseline(latest.I_meas, 5.5),
#                     V_meas=baseline(latest.V_meas, 48.5),
#                     T_meas=baseline(latest.T_meas, 0.66),
#                     SOC_pred=baseline(latest.SOC_pred, 100.66),
#                     Cycle_pred=baseline(latest.Cycle_pred, 10.66),
#                     IR_pred=baseline(latest.IR_pred, 0.01),

#                     created_at=simulated_date,
#                 )

#                 db.add(new_record)
#                 print(f"[{simulated_date}] Inserted new data for {instance_id}")

#             db.commit()
#             day_offset += 1

#             if len(stopped_batteries) == len(instance_ids):
#                 print("All batteries crossed the threshold. Stopping updates.")
#                 break

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)



# async def auto_update_battery_data():
#     """Continuously insert simulated battery data every 5 seconds for all batteries.
#     When risk_score crosses 100, it keeps fluctuating within ±5–10 range (does not stop).
#     """
#     await asyncio.sleep(3)  # Wait for app startup

#     day_offset = 0  # Simulated day counter
#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

#     while True:
#         db: Session = SessionLocal()
#         try:
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]

#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )
#                 if not latest:
#                     continue

#                 simulated_date = start_date + timedelta(days=day_offset)

#                 # Helper to handle first zeros
#                 def baseline(value, base):
#                     return base if value == 0 else vary(value)

#                 # If risk_score crosses 100, oscillate around 100 (± up to 10)
#                 if latest.risk_score is not None and latest.risk_score > 100.0:
#                     new_risk_score = 100 + random.uniform(-10, 10)
#                 elif latest.risk_score is not None and latest.risk_score < 90.0:
#                     new_risk_score = 90 + random.uniform(0, 10)
#                 else:
#                     new_risk_score = vary(latest.risk_score if latest.risk_score else 50.0)

#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     I_meas=baseline(latest.I_meas, 5.5),
#                     V_meas=baseline(latest.V_meas, 48.5),
#                     T_meas=baseline(latest.T_meas, 0.66),
#                     SOC_pred=baseline(latest.SOC_pred, 100.66),
#                     Cycle_pred=baseline(latest.Cycle_pred, 10.66),
#                     IR_pred=baseline(latest.IR_pred, 0.01),
#                     risk_score=new_risk_score,

#                     created_at=simulated_date,
#                 )

#                 db.add(new_record)
#                 print(f"[{simulated_date}] Inserted new data for {instance_id} (risk_score={new_risk_score:.2f})")

#             db.commit()
#             day_offset += 1

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)

# -----------------------------------increasea and decrease around 100 ------------------
# async def auto_update_battery_data():
#     """Continuously insert simulated battery data every 5 seconds for all batteries.
#     When risk_score >= 100, it stops increasing and fluctuates around 100 ± 10.
#     """
#     await asyncio.sleep(3)  # Wait for app startup
#     day_offset = 0
#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

#     while True:
#         db: Session = SessionLocal()
#         try:
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]
#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )
#                 if not latest:
#                     continue

#                 simulated_date = start_date + timedelta(days=day_offset)

#                 def baseline(value, base):
#                     return base if value == 0 else value

#                 def stable(v):
#                     """Keep value roughly stable, add only tiny random noise."""
#                     return v + random.uniform(-0.3, 0.3) if v else v

#                 def small_increase(v):
#                     """Gradually increase the value."""
#                     return v + (v * 0.03) + random.uniform(0.1, 0.5)

#                 # Main decision: if risk_score already high
#                 if latest.risk_score is not None and latest.risk_score >= 100.0:
#                     # Fluctuate near 100 ± 10, not reset randomly
#                     new_risk_score = 100 + random.uniform(-10, 10)

#                     new_record = BatteryLiveData(
#                         instance_id=latest.instance_id,
#                         floor_name=latest.floor_name,
#                         email=latest.email,

#                         I_meas=stable(latest.I_meas),
#                         V_meas=stable(latest.V_meas),
#                         T_meas=stable(latest.T_meas),
#                         SOC_pred=stable(latest.SOC_pred),
#                         Cycle_pred=stable(latest.Cycle_pred),
#                         IR_pred=stable(latest.IR_pred),
#                         risk_score=new_risk_score,

#                         created_at=simulated_date,
#                     )

#                 else:
#                     # Normal growth until risk_score reaches 100
#                     new_risk_score = latest.risk_score + random.uniform(1, 5) if latest.risk_score else 3.0
#                     new_record = BatteryLiveData(
#                         instance_id=latest.instance_id,
#                         floor_name=latest.floor_name,
#                         email=latest.email,

#                         I_meas=small_increase(baseline(latest.I_meas, 5.5)),
#                         V_meas=small_increase(baseline(latest.V_meas, 48.5)),
#                         T_meas=small_increase(baseline(latest.T_meas, 0.66)),
#                         SOC_pred=small_increase(baseline(latest.SOC_pred, 100.66)),
#                         Cycle_pred=small_increase(baseline(latest.Cycle_pred, 10.66)),
#                         IR_pred=baseline(latest.IR_pred, 0.01),
#                         risk_score=new_risk_score,

#                         created_at=simulated_date,
#                     )

#                 db.add(new_record)
#                 print(
#                     f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted data for {instance_id} "
#                     f"(risk_score={new_record.risk_score:.2f})"
#                 )

#             db.commit()
#             day_offset += 1

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)


# ---------------------Auto increment working code ----------------------------------------------
from datetime import datetime, timedelta
async def auto_update_battery_data():
    """Continuously insert simulated battery data every 5 seconds for all batteries.
    When risk_score >= 100, it stops increasing and fluctuates around 100 ± 10.
    """
    await asyncio.sleep(3)  # Wait for app startup
    month_offset = 0
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    while True:
        db: Session = SessionLocal()
        try:
            instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]
            if not instance_ids:
                print("No battery data found to base new records on.")
                await asyncio.sleep(5)
                continue
                

            for instance_id in instance_ids:
                latest = (
                    db.query(BatteryLiveData)
                    .filter(BatteryLiveData.instance_id == instance_id)
                    .order_by(BatteryLiveData.id.desc())
                    .first()
                )
                if not latest:
                    continue

                # simulated_date = start_date + timedelta(days=day_offset)
                simulated_date = start_date + relativedelta(months=month_offset)

                def baseline(value, base):
                    return base if value == 0 else value

                def stable(v):
                    """Keep value roughly stable, add only tiny random noise."""
                    return v + random.uniform(-0.3, 0.3) if v else v

                def small_increase(v):
                    """Gradually increase the value."""
                    return v + (v * 0.03) + random.uniform(0.1, 0.5)

                # Main decision: if risk_score already high
                if latest.risk_score is not None and latest.risk_score >= 100.0:
                    # Fluctuate near 100 ± 10, not reset randomly
                    

                    new_record = BatteryLiveData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I_meas=stable(latest.I_meas),
                        V_meas=stable(latest.V_meas),
                        T_meas=stable(latest.T_meas),
                        SOC_pred=stable(latest.SOC_pred),
                        Cycle_pred=stable(latest.Cycle_pred),
                        IR_pred=stable(latest.IR_pred),
                        created_at=simulated_date
                    )

                else:
                    # Normal growth until risk_score reaches 100
                    # new_risk_score = latest.risk_score + random.uniform(1, 5) if latest.risk_score else 3.0
                    new_record = BatteryLiveData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I_meas=small_increase(baseline(latest.I_meas, 5.5)),
                        V_meas=small_increase(baseline(latest.V_meas, 48.5)),
                        T_meas=small_increase(baseline(latest.T_meas, 0.66)),
                        SOC_pred=small_increase(baseline(latest.SOC_pred, 100.66)),
                        Cycle_pred=small_increase(baseline(latest.Cycle_pred, 10.66)),
                        IR_pred=baseline(latest.IR_pred, 0.01),
                        created_at=simulated_date
                    )

                db.add(new_record)
                print(
                    f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted data for {instance_id}"
                   
                )

            db.commit()
            month_offset += 1

        except Exception as e:
            print(f"[ERROR] auto_update_battery_data failed: {e}")
        finally:
            db.close()

        await asyncio.sleep(5)



#------------------working code stops at 100---------------------------------------------------------------------


# async def auto_update_battery_data():
#     """Continuously insert simulated battery data every 5 seconds for all batteries.
#     Stops inserting once a battery's risk_score reaches or exceeds 100.
#     """
#     await asyncio.sleep(3)  # Wait for app startup
#     month_offset = 0
#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
#     stopped_batteries = set()  # Track batteries that reached/exceeded 100

#     while True:
#         db: Session = SessionLocal()
#         try:
#             instance_ids = [row[0] for row in db.query(BatteryLiveData.instance_id).distinct().all()]
#             if not instance_ids:
#                 print("No battery data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 # Skip batteries already stopped
#                 if instance_id in stopped_batteries:
#                     continue

#                 latest = (
#                     db.query(BatteryLiveData)
#                     .filter(BatteryLiveData.instance_id == instance_id)
#                     .order_by(BatteryLiveData.id.desc())
#                     .first()
#                 )

#                 if not latest:
#                     continue

#                 # Stop updating if risk_score >= 100
#                 if latest.risk_score is not None and latest.risk_score >= 100.0:
#                     print(f"[STOP] {instance_id} risk_score={latest.risk_score:.2f} reached 100.0 — stopped updating.")
#                     stopped_batteries.add(instance_id)
#                     continue

                

#                 def baseline(value, base):
#                     return base if value == 0 else value

#                 def small_increase(v):
#                     """Gradually increase the value."""
#                     return v + (v * 0.03) + random.uniform(0.1, 0.5)

#                 simulated_date = start_date + relativedelta(months=month_offset)
#                 new_record = BatteryLiveData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     I_meas=small_increase(baseline(latest.I_meas, 5.5)),
#                     V_meas=small_increase(baseline(latest.V_meas, 48.5)),
#                     T_meas=small_increase(baseline(latest.T_meas, 0.66)),
#                     SOC_pred=small_increase(baseline(latest.SOC_pred, 100.66)),
#                     Cycle_pred=small_increase(baseline(latest.Cycle_pred, 10.66)),
#                     IR_pred=baseline(latest.IR_pred, 0.01),
#                     created_at=simulated_date
#                 )

#                 db.add(new_record)
#                 print(
#                     f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted data for {instance_id} "
                    
#                 )

#             db.commit()
#             month_offset += 1

#         except Exception as e:
#             print(f"[ERROR] auto_update_battery_data failed: {e}")
#         finally:
#             db.close()
#         await asyncio.sleep(5)


# ----------------------------------------------------------------------------
# async def auto_update_ac_data():
#     """Continuously insert simulated AC data every 5 seconds for all AC instances.
#     When risk_score >= 100, it stops increasing and fluctuates around 100 ± 10.
#     """
#     await asyncio.sleep(3)  # Wait for app startup
#     day_offset = 0
#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

#     while True:
#         db: Session = SessionLocal()
#         try:
#             instance_ids = [row[0] for row in db.query(ACData.instance_id).distinct().all()]
#             if not instance_ids:
#                 print("No AC data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 latest = (
#                     db.query(ACData)
#                     .filter(ACData.instance_id == instance_id)
#                     .order_by(ACData.id.desc())
#                     .first()
#                 )
#                 if not latest:
#                     continue

#                 simulated_date = start_date + timedelta(days=day_offset)

#                 def baseline(value, base):
#                     return base if value == 0 else value

#                 def stable(v):
#                     """Keep value roughly stable, add only tiny random noise."""
#                     return v + random.uniform(-0.3, 0.3) if v else v

#                 def small_increase(v):
#                     """Gradually increase the value."""
#                     return v + (v * 0.03) + random.uniform(0.1, 0.5)

#                 # --- handle risk score logic ---
#                 if latest.risk_score is not None and latest.risk_score >= 100.0:
#                     new_risk_score = 100 + random.uniform(-10, 10)
#                     new_record = ACData(
#                         instance_id=latest.instance_id,
#                         floor_name=latest.floor_name,
#                         email=latest.email,

#                         I_measured=stable(latest.I_measured),
#                         Irated=stable(latest.Irated),
#                         THD=stable(latest.THD),
#                         Tc=stable(latest.Tc),
#                         ESR=stable(latest.ESR),
#                         C_drop=stable(latest.C_drop),
#                         RHpcb=stable(latest.RHpcb),
#                         IR=stable(latest.IR),
#                         Ires=stable(latest.Ires),
#                         risk_score=new_risk_score,

#                         created_at=simulated_date,
#                     )
#                 else:
#                     # --- gradually increase until risk_score reaches 100 ---
#                     new_record = ACData(
#                         instance_id=latest.instance_id,
#                         floor_name=latest.floor_name,
#                         email=latest.email,

#                         I_measured=small_increase(baseline(latest.I_measured, 13)),
#                         Irated=small_increase(baseline(latest.Irated, 15)),
#                         THD=small_increase(baseline(latest.THD, 7)),
#                         Tc=small_increase(baseline(latest.Tc, 80)),
#                         ESR=small_increase(baseline(latest.ESR, 1.1)),
#                         C_drop=small_increase(baseline(latest.C_drop, 0.8)),
#                         RHpcb=small_increase(baseline(latest.RHpcb, 65)),
#                         IR=small_increase(baseline(latest.IR, 2)),
#                         Ires=small_increase(baseline(latest.Ires, 1)),
                       
#                         created_at=simulated_date,
#                     )

#                 db.add(new_record)
#                 print(
#                     f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted AC data for {instance_id} "
#                     # f"(risk_score={new_record.risk_score:.2f})"
#                 )

#             db.commit()
#             day_offset += 1

#         except Exception as e:
#             print(f"[ERROR] auto_update_ac_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)        

# ---------------------------Working ac code ------------------------------------------------------------
async def auto_update_ac_data():
    """Continuously insert simulated AC data every 5 seconds for all AC instances.
    When risk_score >= 80, it stops increasing and fluctuates around 80 ± 10.
    """
    await asyncio.sleep(3)  # Wait for app startup
    month_offset = 0
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    while True:
        db: Session = SessionLocal()
        try:
            instance_ids = [row[0] for row in db.query(ACData.instance_id).distinct().all()]
            if not instance_ids:
                print("No AC data found to base new records on.")
                await asyncio.sleep(5)
                continue

            for instance_id in instance_ids:
                latest = (
                    db.query(ACData)
                    .filter(ACData.instance_id == instance_id)
                    .order_by(ACData.id.desc())
                    .first()
                )
                if not latest:
                    continue

                # simulated_date = start_date + timedelta(months=month_offset)
                simulated_date = start_date + relativedelta(months=month_offset)

                def baseline(value, base):
                    return base if value == 0 else value

                def stable(v):
                    """Keep value roughly stable, add only tiny random noise."""
                    return v + random.uniform(-0.1, 0.1) if v else v

                def small_increase(v):
                    """Gradually increase the value."""
                    return v + (v * 0.01) + random.uniform(0.1, 0.1)

                # --- handle risk score logic ---
                if latest.risk_score is not None and latest.risk_score >= 150.0:
                    # Fluctuate around 80 ± 10
                   
                    new_record = ACData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I_measured=stable(latest.I_measured),
                        Irated=stable(latest.Irated),
                        THD=stable(latest.THD),
                        Tc=stable(latest.Tc),
                        ESR=stable(latest.ESR),
                        C_drop=stable(latest.C_drop),
                        RHpcb=stable(latest.RHpcb),
                        IR=stable(latest.IR),
                        Ires=stable(latest.Ires),
                        created_at=simulated_date
                    )
                else:
                    # --- gradually increase until risk_score reaches 80 ---
                    new_record = ACData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I_measured=small_increase(baseline(latest.I_measured, 13)),
                        Irated=small_increase(baseline(latest.Irated, 15)),
                        THD=small_increase(baseline(latest.THD, 7)),
                        Tc=small_increase(baseline(latest.Tc, 80)),
                        ESR=small_increase(baseline(latest.ESR, 1.1)),
                        C_drop=small_increase(baseline(latest.C_drop, 0.8)),
                        RHpcb=small_increase(baseline(latest.RHpcb, 65)),
                        IR=small_increase(baseline(latest.IR, 2)),
                        Ires=small_increase(baseline(latest.Ires, 1)),
                        created_at=simulated_date
                    )

                db.add(new_record)
                print(
                    f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted AC data for {instance_id} "
                   
                )

            db.commit()
            month_offset += 1

        except Exception as e:
            print(f"[ERROR] auto_update_ac_data failed: {e}")
        finally:
            db.close()

        await asyncio.sleep(5)




# async def auto_update_ac_data():
#     """Continuously insert simulated AC data every 5 seconds for all AC instances.
#     Stops inserting once a AC's risk_score reaches or exceeds 100.
#     """
#     await asyncio.sleep(3)  # Wait for app startup
#     month_offset = 0
#     start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
#     stopped_acs = set()  # Track AC units that reached/exceeded 100

#     while True:
#         db: Session = SessionLocal()
#         try:
#             instance_ids = [row[0] for row in db.query(ACData.instance_id).distinct().all()]
#             if not instance_ids:
#                 print("No AC data found to base new records on.")
#                 break

#             for instance_id in instance_ids:
#                 # Skip ACs already stopped
#                 if instance_id in stopped_acs:
#                     continue

#                 latest = (
#                     db.query(ACData)
#                     .filter(ACData.instance_id == instance_id)
#                     .order_by(ACData.id.desc())
#                     .first()
#                 )
#                 if not latest:
#                     continue

#                 # Stop updating if risk_score >= 100
#                 if latest.risk_score is not None and latest.risk_score >= 150.0:
#                     print(f"[STOP] {instance_id} risk_score={latest.risk_score:.2f} reached 100.0 — stopped updating.")
#                     stopped_acs.add(instance_id)
#                     continue

#                 def baseline(value, base):
#                     return base if value == 0 else value

#                 def small_increase(v):
#                     """Gradually increase the value."""
#                     return v + (v * 0.03) + random.uniform(0.1, 0.5)

#                 simulated_date = start_date + relativedelta(months=month_offset)
#                 new_record = ACData(
#                     instance_id=latest.instance_id,
#                     floor_name=latest.floor_name,
#                     email=latest.email,

#                     I_measured=small_increase(baseline(latest.I_measured, 13)),
#                     Irated=small_increase(baseline(latest.Irated, 15)),
#                     THD=small_increase(baseline(latest.THD, 7)),
#                     Tc=small_increase(baseline(latest.Tc, 80)),
#                     ESR=small_increase(baseline(latest.ESR, 1.1)),
#                     C_drop=small_increase(baseline(latest.C_drop, 0.8)),
#                     RHpcb=small_increase(baseline(latest.RHpcb, 65)),
#                     IR=small_increase(baseline(latest.IR, 2)),
#                     Ires=small_increase(baseline(latest.Ires, 1)),

#                     created_at=simulated_date,
#                 )

#                 db.add(new_record)
#                 print(
#                     f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted AC data for {instance_id}"
#                 )

#             db.commit()
#             month_offset += 1

#         except Exception as e:
#             print(f"[ERROR] auto_update_ac_data failed: {e}")
#         finally:
#             db.close()

#         await asyncio.sleep(5)



# -------------------------------------------------------------------------------------------


async def auto_update_ups_data():
    """Continuously insert simulated UPS data every 5 seconds for all UPS instances.
    When risk_score >= 150, it stops increasing and fluctuates around stable values.
    """
    await asyncio.sleep(3)
    month_offset = 0
    start_date = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    while True:
        db: Session = SessionLocal()
        try:
            instance_ids = [row[0] for row in db.query(UPSData.instance_id).distinct().all()]
            if not instance_ids:
                print("No UPS data found to base new records on.")
                await asyncio.sleep(5)
                continue

            for instance_id in instance_ids:
                latest = (
                    db.query(UPSData)
                    .filter(UPSData.instance_id == instance_id)
                    .order_by(UPSData.id.desc())
                    .first()
                )
                if not latest:
                    continue

                simulated_date = start_date + relativedelta(months=month_offset)

                # --- helper functions ---
                def baseline(value, base):
                    return base if value == 0 else value

                def stable(v):
                    """Keep value roughly stable, add small random noise."""
                    return v + random.uniform(-0.2, 0.2) if v else v

                def strong_increase(v, step):
                    """Increase by a stronger fixed step + small noise."""
                    return v + step + random.uniform(-0.3, 0.3)

                # --- Risk score logic ---
                if latest.risk_score is not None and latest.risk_score >= 150.0:
                    # Fluctuate around stable values
                    new_record = UPSData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I=stable(latest.I),
                        V=stable(latest.V),
                        PF=stable(latest.PF),
                        TUPS=stable(latest.TUPS),
                        IL=stable(latest.IL),
                        THD=stable(latest.THD),
                        created_at=simulated_date
                    )
                else:
                    # Gradually increase values in noticeable steps
                    new_record = UPSData(
                        instance_id=latest.instance_id,
                        floor_name=latest.floor_name,
                        email=latest.email,

                        I=strong_increase(baseline(latest.I, 10), 5),    # +5 each cycle
                        V=strong_increase(baseline(latest.V, 230), 5),   # +5 each cycle
                        PF=min(strong_increase(baseline(latest.PF, 0.97), 0.05), 1.5),  # cap for realism
                        TUPS=strong_increase(baseline(latest.TUPS, 10), 2),  # +2 each cycle
                        IL=strong_increase(baseline(latest.IL, 4), 0.5),     # +0.5 each cycle
                        THD=strong_increase(baseline(latest.THD, 2.5), 0.3), # +0.3 each cycle
                        created_at=simulated_date
                    )

                db.add(new_record)
                print(f"[{simulated_date.strftime('%Y-%m-%d')}] Inserted UPS data for {instance_id}")

            db.commit()
            month_offset += 1

        except Exception as e:
            print(f"[ERROR] auto_update_ups_data failed: {e}")
        finally:
            db.close()

        await asyncio.sleep(5)



#--------------------------- BACKGROUND TASK ---------------------------
async def run_all_risk_calculations_periodically():
    """Run all risk APIs every 5 seconds automatically."""
    await asyncio.sleep(3)  # wait until app fully starts

    while True:
        try:
            print(f"\n[{datetime.utcnow()}] --- Running all risk calculations ---")
            db = SessionLocal()

            # Call each function directly (no HTTP)
            # calculate_all_ac_risks(db)
            # calculate_battery_risk_all_api(db)
            # calculate_all_ups_risk(db)
            # calculate_risk_all(db)
            # calculate_wiring_risk_all()
            # calculate_switchboard_risk_api(db)

            db.close()
            print(f"[{datetime.utcnow()}] --- Risk calculations done ---\n")

        except Exception as e:
            print(f"[ERROR] Background risk calculation failed: {e}")

        await asyncio.sleep(5)  # wait 5 seconds before next run


# #---------------- START ON SERVER LAUNCH ----------------------------------
# @app.on_event("startup")
# async def start_background_tasks():
#     asyncio.create_task(run_all_risk_calculations_periodically())
#     # asyncio.create_task(auto_update_battery_data())
#     # asyncio.create_task(auto_update_ac_data())
#     asyncio.create_task(auto_update_ups_data())


# @app.post("/upload_excel", response_model=schemas.MessageResponse)
# async def upload_excel(
#     user_name: str = Form(...),
#     file: UploadFile = File(...),
#     db: Session = Depends(get_db)
# ):
#     try:
#         # Read Excel into DataFrame
#         contents = await file.read()
#         df = pd.read_excel(BytesIO(contents))

#         # Required columns check
#         required_columns = [
#             "floor_name", "component_name", "instance_id",
#             "position_x", "position_y", "grid_number"
#         ]
#         for col in required_columns:
#             if col not in df.columns:
#                 raise HTTPException(status_code=400, detail=f"Missing column: {col}")

#         # Ensure user exists
#         user_exists = db.query(models.User).filter(models.User.email == user_name).first()
#         if not user_exists:
#             raise HTTPException(status_code=404, detail="User not found")

#         # Convert Excel rows to ComponentBase objects
#         components = []
#         for _, row in df.iterrows():
#             try:
#                 comp = schemas.ComponentBase(
#                     floor_name=str(row["floor_name"]).strip(),
#                     component_name=str(row["component_name"]).strip(),
#                     instance_id=str(row["instance_id"]).strip(),
#                     position_x=float(row["position_x"]),
#                     position_y=float(row["position_y"]),
#                     grid_number=int(row["grid_number"])
#                 )
#                 components.append(comp)
#             except Exception as e:
#                 raise HTTPException(status_code=400, detail=f"Invalid row format: {e}")

#         # Reuse same logic as save_components
#         if not components:
#             raise HTTPException(status_code=400, detail="No valid component data found in Excel")

#         floor_names = {comp.floor_name for comp in components}

#         for floor_name in floor_names:
#             existing_components = db.query(models.ComponentLayout).filter(
#                 models.ComponentLayout.user_name == user_name,
#                 models.ComponentLayout.floor_name == floor_name
#             ).all()

#             existing_keys = {(comp.floor_name, comp.instance_id) for comp in existing_components}
#             new_keys = {(comp.floor_name, comp.instance_id) for comp in components if comp.floor_name == floor_name}

#             # DELETE missing components
#             for comp in existing_components:
#                 if (comp.floor_name, comp.instance_id) not in new_keys:
#                     component_type = comp.component_name.lower()
#                     db.delete(comp)

#                     # Delete linked data from equipment tables
#                     if component_type == "wiring":
#                         item = db.query(models.WiringEquipmentData).filter(
#                             models.WiringEquipmentData.email == user_name,
#                             models.WiringEquipmentData.floor_name == comp.floor_name,
#                             models.WiringEquipmentData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#                     elif component_type == "panel":
#                         item = db.query(models.PanelRiskData).filter(
#                             models.PanelRiskData.email == user_name,
#                             models.PanelRiskData.floor_name == comp.floor_name,
#                             models.PanelRiskData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#                     elif component_type == "ac":
#                         item = db.query(models.ACData).filter(
#                             models.ACData.email == user_name,
#                             models.ACData.floor_name == comp.floor_name,
#                             models.ACData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#                     elif component_type == "battery":
#                         item = db.query(models.BatteryLiveData).filter(
#                             models.BatteryLiveData.email == user_name,
#                             models.BatteryLiveData.floor_name == comp.floor_name,
#                             models.BatteryLiveData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#                     elif component_type == "ups":
#                         item = db.query(models.UPSData).filter(
#                             models.UPSData.email == user_name,
#                             models.UPSData.floor_name == comp.floor_name,
#                             models.UPSData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#                     elif component_type == "switchboard":
#                         item = db.query(models.SwitchboardLiveData).filter(
#                             models.SwitchboardLiveData.email == user_name,
#                             models.SwitchboardLiveData.floor_name == comp.floor_name,
#                             models.SwitchboardLiveData.instance_id == comp.instance_id
#                         ).first()
#                         if item: db.delete(item)

#             # INSERT/UPDATE per floor
#             for comp in components:
#                 if comp.floor_name != floor_name:
#                     continue

#                 if not comp.component_name or not comp.instance_id:
#                     continue

#                 # --- ComponentLayout ---
#                 existing_component = db.query(models.ComponentLayout).filter(
#                     models.ComponentLayout.user_name == user_name,
#                     models.ComponentLayout.floor_name == comp.floor_name,
#                     models.ComponentLayout.instance_id == comp.instance_id
#                 ).first()

#                 if existing_component:
#                     existing_component.position_x = comp.position_x
#                     existing_component.position_y = comp.position_y
#                     existing_component.component_name = comp.component_name
#                     existing_component.grid_number = comp.grid_number
#                 else:
#                     new_component = models.ComponentLayout(
#                         user_name=user_name,
#                         floor_name=comp.floor_name,
#                         component_name=comp.component_name,
#                         instance_id=comp.instance_id,
#                         position_x=comp.position_x,
#                         position_y=comp.position_y,
#                         grid_number=comp.grid_number
#                     )
#                     db.add(new_component)

#                 # --- Sync equipment tables ---
#                 ctype = comp.component_name.lower()

#                 if ctype == "wiring":
#                     if not db.query(models.WiringEquipmentData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.WiringEquipmentData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id))

#                 elif ctype == "panel":
#                     if not db.query(models.PanelRiskData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.PanelRiskData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id))

#                 elif ctype == "ac":
#                     if not db.query(models.ACData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.ACData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id, created_at=datetime.now()))

#                 elif ctype == "battery":
#                     if not db.query(models.BatteryLiveData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.BatteryLiveData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id, created_at=datetime.now()))

#                 elif ctype == "ups":
#                     if not db.query(models.UPSData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.UPSData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id))

#                 elif ctype == "switchboard":
#                     if not db.query(models.SwitchboardLiveData).filter_by(
#                         email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
#                     ).first():
#                         db.add(models.SwitchboardLiveData(email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id))

#         db.commit()

#         return {
#             "message": "Excel uploaded and components saved successfully!",
#             "rows_inserted": len(df),
#             "user_name": user_name,
#             "columns": list(df.columns)
#         }

#     except HTTPException:
#         raise
#     except Exception as e:
#         db.rollback()
#         raise HTTPException(status_code=400, detail=f"Error processing Excel: {e}")



# -------------------------------------------------------------------------------------------



@app.post("/upload_excel", response_model=schemas.MessageResponse)
async def upload_excel(
    user_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        # Read Excel into DataFrame
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))


        # ---------------- CLEANUP ----------------

        # 1. Remove unwanted extra blank columns (Unnamed: x)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]

        # 2. Remove rows where all required fields are empty
        df = df.dropna(
            subset=["floor_name", "component_name", "instance_id", "grid_number", "location"],
            how="all"
        )

        # 3. Remove rows where grid_number is NaN
        df = df[df["grid_number"].notna()]

        # Remove completely empty rows (Excel often has hidden blank rows)
        df = df.dropna(how="all")

        # Updated required columns (position_x & position_y removed, location added)
        required_columns = [
            "floor_name", "component_name", "instance_id",
            "grid_number", "location"
        ]
        for col in required_columns:
            if col not in df.columns:
                raise HTTPException(status_code=400, detail=f"Missing column: {col}")

        # Ensure user exists
        user_exists = db.query(models.User).filter(models.User.email == user_name).first()
        if not user_exists:
            raise HTTPException(status_code=404, detail="User not found")

        # Convert Excel rows to ComponentBase objects
        components = []
        for index, row in df.iterrows():
            try:
                grid_val = row["grid_number"]

                # Debug: raise detailed info before conversion
                if pd.isna(grid_val):
                    raise HTTPException(
                        status_code=400,
                        detail=f"Row {index+2}: grid_number is NaN. Full row = {row.to_dict()}"
                    )

                comp = schemas.ExcelComponentBase(
                    floor_name=str(row["floor_name"]).strip(),
                    component_name=str(row["component_name"]).strip(),
                    instance_id=str(row["instance_id"]).strip(),
                    grid_number=int(grid_val),
                    location=str(row["location"]).strip(),
                )

                components.append(comp)

            except Exception as e:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid row format at Excel row {index+2}: {e}. Row = {row.to_dict()}"
                )


        if not components:
            raise HTTPException(status_code=400, detail="No valid component data found in Excel")

        floor_names = {comp.floor_name for comp in components}

        # ---------- DELETE MISSING COMPONENTS ----------
        for floor_name in floor_names:
            existing_components = db.query(models.ComponentLayout).filter(
                models.ComponentLayout.user_name == user_name,
                models.ComponentLayout.floor_name == floor_name
            ).all()

            existing_keys = {(comp.floor_name, comp.instance_id) for comp in existing_components}
            new_keys = {(comp.floor_name, comp.instance_id) for comp in components if comp.floor_name == floor_name}

            for comp in existing_components:
                if (comp.floor_name, comp.instance_id) not in new_keys:
                    component_type = comp.component_name.lower()
                    db.delete(comp)

                    # delete linked equipment table entry
                    if component_type == "wiring":
                        item = db.query(models.WiringEquipmentData).filter(
                            models.WiringEquipmentData.email == user_name,
                            models.WiringEquipmentData.floor_name == comp.floor_name,
                            models.WiringEquipmentData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

                    elif component_type == "panel":
                        item = db.query(models.PanelRiskData).filter(
                            models.PanelRiskData.email == user_name,
                            models.PanelRiskData.floor_name == comp.floor_name,
                            models.PanelRiskData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

                    elif component_type == "ac":
                        item = db.query(models.ACData).filter(
                            models.ACData.email == user_name,
                            models.ACData.floor_name == comp.floor_name,
                            models.ACData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

                    elif component_type == "battery":
                        item = db.query(models.BatteryLiveData).filter(
                            models.BatteryLiveData.email == user_name,
                            models.BatteryLiveData.floor_name == comp.floor_name,
                            models.BatteryLiveData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

                    elif component_type == "ups":
                        item = db.query(models.UPSData).filter(
                            models.UPSData.email == user_name,
                            models.UPSData.floor_name == comp.floor_name,
                            models.UPSData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

                    elif component_type == "switchboard":
                        item = db.query(models.SwitchboardLiveData).filter(
                            models.SwitchboardLiveData.email == user_name,
                            models.SwitchboardLiveData.floor_name == comp.floor_name,
                            models.SwitchboardLiveData.instance_id == comp.instance_id
                        ).first()
                        if item: db.delete(item)

        # ---------- INSERT / UPDATE ----------
        for comp in components:
            # Check existing component layout
            existing_component = db.query(models.ComponentLayout).filter(
                models.ComponentLayout.user_name == user_name,
                models.ComponentLayout.floor_name == comp.floor_name,
                models.ComponentLayout.instance_id == comp.instance_id
            ).first()

            if existing_component:
                # Update only the fields that come from Excel
                existing_component.component_name = comp.component_name
                # existing_component.grid_number = comp.grid_number
                existing_component.location = comp.location

                # ❌ DO NOT reset position_x and position_y
                # existing_component.position_x = 0
                # existing_component.position_y = 0

            else:
                # Only new rows get default 0,0
                new_component = models.ComponentLayout(
                    user_name=user_name,
                    floor_name=comp.floor_name,
                    component_name=comp.component_name,
                    instance_id=comp.instance_id,
                    grid_number=comp.grid_number,
                    location=comp.location,
                    position_x=0,   # allowed for NEW
                    position_y=0
                )
                db.add(new_component)


            # ---------- SYNC EQUIPMENT TABLES ----------
            ctype = comp.component_name.lower()

            if ctype == "wiring":
                if not db.query(models.WiringEquipmentData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.WiringEquipmentData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif ctype == "panel":
                if not db.query(models.PanelRiskData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.PanelRiskData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif ctype == "ac":
                if not db.query(models.ACData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.ACData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id, created_at=datetime.now()
                    ))

            elif ctype == "battery":
                if not db.query(models.BatteryLiveData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.BatteryLiveData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id, created_at=datetime.now()
                    ))

            elif ctype == "ups":
                if not db.query(models.UPSData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.UPSData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

            elif ctype == "switchboard":
                if not db.query(models.SwitchboardLiveData).filter_by(
                    email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                ).first():
                    db.add(models.SwitchboardLiveData(
                        email=user_name, floor_name=comp.floor_name, instance_id=comp.instance_id
                    ))

        db.commit()

        return {
            "message": "Excel uploaded and components saved successfully!",
            "rows_inserted": len(df),
            "user_name": user_name,
            "columns": list(df.columns)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error processing Excel: {e}")





# @app.post("/company/users", response_model=List[dict])
# def get_users_by_company(payload: CompanyRequest):
#     try:
#         company_name = payload.company_name

#         with engine.connect() as conn:
#             query = """
#                 SELECT name, email, contact_number, role 
#                 FROM users 
#                 WHERE company_name = :company_name
#             """

#             result = conn.execute(
#                 text(query),
#                 {"company_name": company_name}
#             )

#             users = [dict(row._mapping) for row in result]

#             if not users:
#                 raise HTTPException(
#                     status_code=404,
#                     detail=f"No users found for company: {company_name}"
#                 )

#             return users

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=str(e))



@app.post("/company/users", response_model=List[dict])
def get_users_by_company(payload: CompanyRequest):
    try:
        company_name = payload.company_name

        with engine.connect() as conn:
            query = """
                SELECT name, email, contact_number, approval_status, role 
                FROM users 
                WHERE company_name = :company_name
                AND role = 'user'
            """

            result = conn.execute(
                text(query),
                {"company_name": company_name}
            )

            users = [dict(row._mapping) for row in result]

            if not users:
                raise HTTPException(
                    status_code=404,
                    detail=f"No users with role 'user' found for company: {company_name}"
                )

            return users

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@app.put("/users/update-approval")
async def update_approval_status(status_update: ApprovalStatusUpdate, db: Session = Depends(get_db)):
    
    # Find user by email
    user = db.query(User).filter(User.email == status_update.email).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found with given email")

    # Update status
    user.approval_status = status_update.approval_status
    db.commit()
    db.refresh(user)

    # Create email message body
    message = MessageSchema(
        subject="Account Approval Status Updated",
        recipients=[user.email],
        body=(
            f"Hello {user.name},\n\n"
            f"Your account approval status has been updated to: {user.approval_status.upper()}.\n\n"
            f"Thank you,\nMepstra IT Solutions"
        ),
        subtype="plain",
    )

    # Send email
    fm = FastMail(conf)
    await fm.send_message(message)

    return {
        "message": "Approval status updated & email sent successfully",
        "email": user.email,
        "new_status": user.approval_status
    }  


@app.post("/save_token_details")
async def save_token_details(request: SaveTokenRequest, db: Session = Depends(get_db)):

    # Check user exists
    user = db.query(models.User).filter(models.User.email == request.email).first()

    if not user:
        raise HTTPException(
            status_code=404,
            detail="User not found"
        )

    # Save / update token number
    user.token_number = request.token_number
    db.commit()
    db.refresh(user)

    return {
        "message": "Token saved successfully",
        "email": user.email,
        "token_number": user.token_number
    }

# DATABASE_URL = "mysql+pymysql://root:root@localhost:3306/fire_prediction_db"

# engine = create_engine(DATABASE_URL)
# metadata = MetaData()
# metadata.reflect(bind=engine)

# component_layouts = metadata.tables['component_layouts']

# # Request body model
# class UserRequest(BaseModel):
#     user_name: str

# @app.post("/export_equipment/")
# def export_equipment(request: UserRequest):
#     user_name = request.user_name
    
#     # Fetch data for the user
#     query = select(
#         component_layouts.c.floor_name,
#         component_layouts.c.component_name,
#         component_layouts.c.instance_id,
#         component_layouts.c.grid_number,
#         component_layouts.c.location
#     ).where(component_layouts.c.user_name == user_name)
    
#     with engine.connect() as conn:
#         result = conn.execute(query)
#         rows = result.fetchall()
    
#     if not rows:
#         raise HTTPException(status_code=404, detail="No data found for this user")
    
#     # Convert to DataFrame
#     df = pd.DataFrame(rows, columns=['floor_name', 'component_name', 'instance_id', 'grid_number', 'location'])
    
#     # Generate Excel file name
#     today = datetime.now().strftime("%Y-%m-%d")
#     file_name = f"{user_name}-{today}.xlsx"
    
#     # Save Excel
#     df.to_excel(file_name, index=False)
    
#     # Return file as download
#     return FileResponse(path=file_name, filename=file_name, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')        


# -------------------------------download excel working code-----------------------------------------

# # Database setup
# DATABASE_URL = "mysql+pymysql://root:root@localhost:3306/fire_prediction_db"
# engine = create_engine(DATABASE_URL)
# metadata = MetaData()
# metadata.reflect(bind=engine)
# component_layouts = metadata.tables['component_layouts']

# # Request body model
# class UserRequest(BaseModel):
#     user_name: str

# @app.post("/export_equipment/")
# def export_equipment(request: UserRequest):
#     user_name = request.user_name

#     # Fetch data for the user
#     query = select(
#         component_layouts.c.floor_name,
#         component_layouts.c.component_name,
#         component_layouts.c.instance_id,
#         component_layouts.c.grid_number,
#         component_layouts.c.location
#     ).where(component_layouts.c.user_name == user_name)

#     with engine.connect() as conn:
#         result = conn.execute(query)
#         rows = result.fetchall()

#     if not rows:
#         raise HTTPException(status_code=404, detail="No data found for this user")

#     # Use relative path for template inside project directory
#     project_dir = os.path.dirname(os.path.abspath(__file__))
#     template_path = os.path.join(project_dir, "Equipments_details_file", "components_template.xlsx")

#     wb = load_workbook(template_path)
#     ws = wb.active

#     # Clear existing data starting from row 2
#     if ws.max_row > 1:
#         ws.delete_rows(2, ws.max_row - 1)

#     # Write fetched data to template starting from row 2
#     for i, row in enumerate(rows, start=2):
#         ws.cell(row=i, column=1, value=row.floor_name)
#         ws.cell(row=i, column=2, value=row.component_name)
#         ws.cell(row=i, column=3, value=row.instance_id)
#         ws.cell(row=i, column=4, value=row.grid_number)
#         ws.cell(row=i, column=5, value=row.location)

#     # Generate file name and save in the same folder
#     today = datetime.now().strftime("%Y-%m-%d")
#     file_name = f"{user_name}-{today}.xlsx"
#     save_path = os.path.join(project_dir, "Equipments_details_file", file_name)
#     wb.save(save_path)

#     return FileResponse(
#         path=save_path,
#         filename=file_name,
#         media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )




# Database URL
DATABASE_URL = "mysql+pymysql://root:root@localhost:3306/fire_prediction_db"
engine = create_engine(DATABASE_URL)
metadata = MetaData()
metadata.reflect(bind=engine)
component_layouts = metadata.tables['component_layouts']

# Request body model
class UserRequest(BaseModel):
    user_name: str

@app.post("/export_equipment/")
def export_equipment(request: UserRequest):
    user_name = request.user_name

    # Fetch data for the user
    query = select(
        component_layouts.c.floor_name,
        component_layouts.c.component_name,
        component_layouts.c.instance_id,
        component_layouts.c.grid_number,
        component_layouts.c.location
    ).where(component_layouts.c.user_name == user_name)

    with engine.connect() as conn:
        result = conn.execute(query)
        rows = result.fetchall()

    if not rows:
        raise HTTPException(status_code=404, detail="No data found for this user")

    # Paths
    project_dir = Path(__file__).parent
    template_path = project_dir / "Equipments_details_file" / "components_template.xlsx"
    today = datetime.now().strftime("%Y-%m-%d")
    file_name = f"{user_name}-{today}.xlsx"
    save_path = project_dir / "Equipments_details_file" / file_name

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Unprotect the sheet if protected
    ws.protection.sheet = False

    # Clear existing data from row 2
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Write data to sheet and unlock all cells
    for i, row in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=row.floor_name).protection = None
        ws.cell(row=i, column=2, value=row.component_name).protection = None
        ws.cell(row=i, column=3, value=row.instance_id).protection = None
        ws.cell(row=i, column=4, value=row.grid_number).protection = None
        ws.cell(row=i, column=5, value=row.location).protection = None

    #Do NOT protect the sheet, leave it editable
    # Data validation in template for grid_number remains intact

    wb.save(save_path)

    return FileResponse(
        path=save_path,
        filename=file_name,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


